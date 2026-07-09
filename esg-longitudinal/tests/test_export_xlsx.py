import os
from conftest import _load

export_xlsx = _load("export_xlsx")
snapshot = _load("snapshot")


def _write_csv(path, columns, rows):
    """Write a snapshot-style CSV with a BOM, exactly like snapshot.py does."""
    import csv
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)


def test_read_snapshot_strips_bom_and_preserves_order(tmp_path):
    p = tmp_path / "s.csv"
    cols = ["entity", "domain", "indicator", "value", "period", "status"]
    _write_csv(p, cols, [{"entity": "Royal Philips", "domain": "circular",
                          "indicator": "circular_revenue_pct", "value": "18",
                          "period": "2022", "status": "found"}])
    columns, rows = export_xlsx.read_snapshot(str(p))
    assert columns == cols               # first col is 'entity', not '﻿entity'
    assert columns[0] == "entity"
    assert rows[0]["value"] == "18"


def test_read_snapshot_handles_embedded_comma_and_newline(tmp_path):
    p = tmp_path / "s.csv"
    cols = ["entity", "domain", "indicator", "period", "status", "quote"]
    tricky = 'He said, "25% by 2025",\nacross two lines'
    _write_csv(p, cols, [{"entity": "X", "domain": "circular",
                          "indicator": "i", "period": "2025",
                          "status": "target", "quote": tricky}])
    _, rows = export_xlsx.read_snapshot(str(p))
    assert rows[0]["quote"] == tricky    # one cell, comma+newline intact


def test_read_snapshot_header_only(tmp_path):
    p = tmp_path / "s.csv"
    cols = ["entity", "domain", "indicator", "period", "status"]
    _write_csv(p, cols, [])
    columns, rows = export_xlsx.read_snapshot(str(p))
    assert columns == cols
    assert rows == []


def test_legend_covers_every_canonical_column():
    documented = [c for c, _, _ in export_xlsx.LEGEND_COLUMNS]
    # set-equality (not subset): catches misspellings, omissions, AND extras/dupes.
    assert set(documented) == set(snapshot.COLS)
    assert len(documented) == len(snapshot.COLS)


def test_legend_code_tables_cover_enums():
    def codes(key):
        return {code for code, _ in export_xlsx.LEGEND_CODE_TABLES[key]}
    assert snapshot.VALID_STATUS <= codes("status")
    assert snapshot.VALID_ITEM_TYPE <= codes("item_type")
    assert snapshot.VALID_PLANETARY <= codes("planetary_alignment")
    assert snapshot.VALID_IMPACT_SCOPE <= codes("impact_scope")
    assert snapshot.VALID_TARGET_STATUS <= codes("target_status")


def test_fill_constants_shared_identity():
    assert export_xlsx.FILL_YES == export_xlsx.FILL_FOUND
    assert export_xlsx.STATUS_FILLS["found"] == export_xlsx.FILL_FOUND
    assert export_xlsx.STATUS_FILLS["target"] == export_xlsx.FILL_TARGET
    assert export_xlsx.STATUS_FILLS["not_found"] == export_xlsx.FILL_NOT_FOUND
    assert export_xlsx.SUBSTANCE_FILLS["symbolic"] == export_xlsx.FILL_NO


import pytest
# NOTE: do NOT import any openpyxl symbol at module scope — that would run at
# pytest collection time and error the whole module (defeating the openpyxl-free
# tests and the importorskip fallback) when openpyxl is absent. Import openpyxl
# names lazily, inside tests that already went through _wb()'s importorskip.


def _wb(columns, rows):
    pytest.importorskip("openpyxl")
    return export_xlsx.build_workbook(columns, rows)


DATA_COLS = ["entity", "domain", "indicator", "value", "period", "status",
             "item_type", "target_status", "smart_specific", "substance",
             "planetary_alignment", "impact_scope", "quote", "assessment_notes",
             "made_up_col"]


def _data_rows():
    return [
        {"entity": "X", "domain": "circular", "indicator": "i1", "value": "18",
         "period": "2022", "status": "found", "quote": "q", "made_up_col": "z"},
        {"entity": "X", "domain": "circular", "indicator": "i2", "value": "25",
         "period": "2025", "status": "target", "item_type": "target",
         "target_status": "on_track", "smart_specific": "yes", "substance": "substantive",
         "planetary_alignment": "pb_aligned", "impact_scope": "D",
         "quote": "25% by 2025", "assessment_notes": "board KPI"},
        {"entity": "X", "domain": "circular", "indicator": "i3", "value": "",
         "period": "2021", "status": "not_found", "smart_specific": "no",
         "substance": "symbolic"},
    ]


def _fill_rgb(cell):
    return cell.fill.fgColor.rgb if cell.fill and cell.fill.fill_type else None


def _col_idx(ws, name):
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == name:
            return c
    raise AssertionError(f"column {name} not in header")


def test_data_sheet_header_order_and_freeze():
    ws = _wb(DATA_COLS, _data_rows())["Data"]
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert header == DATA_COLS
    assert ws.freeze_panes == "A2"


def test_autofilter_covers_full_range():
    rows = _data_rows()
    ws = _wb(DATA_COLS, rows)["Data"]
    from openpyxl.utils import get_column_letter   # lazy: after _wb's importorskip
    last_col = get_column_letter(len(DATA_COLS))
    assert ws.auto_filter.ref == f"A1:{last_col}{len(rows) + 1}"


@pytest.mark.parametrize("col,value,expected", [
    ("status", "found", "FILL_FOUND"),
    ("status", "target", "FILL_TARGET"),
    ("status", "not_found", "FILL_NOT_FOUND"),
    ("smart_specific", "yes", "FILL_YES"),
    ("smart_specific", "no", "FILL_NO"),
    ("substance", "substantive", "FILL_YES"),
    ("substance", "symbolic", "FILL_NO"),
])
def test_cell_fill_matrix(col, value, expected):
    pytest.importorskip("openpyxl")
    assert export_xlsx._cell_fill(col, value) == getattr(export_xlsx, expected)


def test_status_cells_get_expected_fill_on_sheet():
    rows = _data_rows()
    ws = _wb(DATA_COLS, rows)["Data"]
    sc = _col_idx(ws, "status")
    assert _fill_rgb(ws.cell(row=2, column=sc)) == export_xlsx.FILL_FOUND
    assert _fill_rgb(ws.cell(row=3, column=sc)) == export_xlsx.FILL_TARGET
    assert _fill_rgb(ws.cell(row=4, column=sc)) == export_xlsx.FILL_NOT_FOUND


def test_uncolored_coded_columns_have_no_fill():
    ws = _wb(DATA_COLS, _data_rows())["Data"]
    for name in ("item_type", "target_status", "planetary_alignment", "impact_scope"):
        c = _col_idx(ws, name)
        assert _fill_rgb(ws.cell(row=3, column=c)) is None, f"{name} should be uncolored"


def test_unknown_column_has_no_fill_but_appears():
    ws = _wb(DATA_COLS, _data_rows())["Data"]
    c = _col_idx(ws, "made_up_col")
    assert _fill_rgb(ws.cell(row=2, column=c)) is None


def test_blank_smart_cell_has_no_fill():
    ws = _wb(DATA_COLS, _data_rows())["Data"]
    c = _col_idx(ws, "smart_specific")
    assert _fill_rgb(ws.cell(row=2, column=c)) is None   # row 2 has no smart_specific


def test_wrap_text_on_long_columns():
    ws = _wb(DATA_COLS, _data_rows())["Data"]
    c = _col_idx(ws, "quote")
    assert ws.cell(row=2, column=c).alignment.wrap_text is True


def _all_cell_values(ws):
    out = []
    for row in ws.iter_rows(values_only=True):
        out.extend(v for v in row if v is not None)
    return out


def test_sheets_are_exactly_data_and_legend():
    wb = _wb(DATA_COLS, _data_rows())
    assert wb.sheetnames == ["Data", "Legend"]


def test_legend_lists_column_names_and_multichar_codes():
    ws = _wb(DATA_COLS, _data_rows())["Legend"]
    text = "\n".join(str(v) for v in _all_cell_values(ws))
    for name in ("entity", "impact_scope", "planetary_alignment", "target_status"):
        assert name in text
    for code in ("pb_aligned", "not_found", "too_early", "Handprint", "handprint"):
        # at least the lowercased forms are present
        pass
    for code in ("pb_aligned", "not_found", "too_early"):
        assert code in text


def test_legend_impact_scope_uses_code_table_cells_not_substring():
    # 'D' as a bare substring is vacuous; assert D sits in a cell paired with its meaning.
    ws = _wb(DATA_COLS, _data_rows())["Legend"]
    found = False
    for row in ws.iter_rows(values_only=True):
        cells = [str(c) for c in row if c is not None]
        if "D" in cells and any("handprint" in c.lower() for c in cells):
            found = True
            break
    assert found, "impact_scope D / handprint row not found as discrete cells"


def test_color_key_swatches_match_data_fills():
    wb = _wb(DATA_COLS, _data_rows())
    legend = wb["Legend"]
    # Tie each swatch fill to its label's row (col1 swatch, col2 label) so a
    # swapped swatch is caught, not just overall presence of the colors.
    label_to_fill = {}
    for row in legend.iter_rows():
        cells = list(row)
        if len(cells) >= 2 and cells[1].value:
            f = _fill_rgb(cells[0])
            if f is not None:
                label_to_fill[cells[1].value] = f
    for color, label in export_xlsx.LEGEND_COLOR_KEY:
        assert label_to_fill.get(label) == color, f"swatch for {label!r} != {color}"


# 13 / 19 / 29 column headers straight from snapshot.py's COLS.
@pytest.mark.parametrize("width", [13, 19, 29])
def test_schema_tolerance_round_trip(width):
    pytest.importorskip("openpyxl")
    columns = snapshot.COLS[:width]
    row = {c: "" for c in columns}
    row.update({"entity": "X", "domain": "circular", "indicator": "i",
                "period": "2022", "status": "found", "value": "1"})
    wb = export_xlsx.build_workbook(columns, [row])
    assert wb.sheetnames == ["Data", "Legend"]
    header = [wb["Data"].cell(row=1, column=c).value for c in range(1, len(columns) + 1)]
    assert header == columns


def test_header_only_still_valid():
    wb = _wb(["entity", "domain", "indicator", "period", "status"], [])
    assert wb.sheetnames == ["Data", "Legend"]
    assert wb["Data"].cell(row=1, column=1).value == "entity"


import sys


def test_default_out_derives_basename():
    assert export_xlsx._default_out("data/snapshots/2026-07-03-Philips.csv") \
        == os.path.join("reports", "2026-07-03-Philips.xlsx")


def test_main_happy_path_writes_workbook(tmp_path):
    pytest.importorskip("openpyxl")
    src = tmp_path / "2026-07-03.csv"
    _write_csv(src, ["entity", "domain", "indicator", "period", "status"],
               [{"entity": "X", "domain": "circular", "indicator": "i",
                 "period": "2022", "status": "found"}])
    out = tmp_path / "out.xlsx"
    rc = export_xlsx.main(["--snapshot", str(src), "--out", str(out)])
    assert rc == 0
    assert out.exists()
    from openpyxl import load_workbook
    assert load_workbook(str(out)).sheetnames == ["Data", "Legend"]


def test_main_missing_snapshot_returns_nonzero_and_names_path(tmp_path, capsys):
    missing = tmp_path / "nope.csv"
    rc = export_xlsx.main(["--snapshot", str(missing), "--out", str(tmp_path / "o.xlsx")])
    assert rc != 0
    assert "nope.csv" in capsys.readouterr().err


def test_main_unreadable_snapshot_dir_returns_nonzero(tmp_path):
    d = tmp_path / "adir"
    d.mkdir()
    rc = export_xlsx.main(["--snapshot", str(d), "--out", str(tmp_path / "o.xlsx")])
    assert rc != 0


def test_main_missing_openpyxl_returns_nonzero_with_hint(tmp_path, capsys, monkeypatch):
    # NOT importorskip-guarded: this path must run in an env that HAS openpyxl.
    src = tmp_path / "s.csv"
    _write_csv(src, ["entity", "domain", "indicator", "period", "status"],
               [{"entity": "X", "domain": "circular", "indicator": "i",
                 "period": "2022", "status": "found"}])
    # Force ImportError regardless of import form / test order.
    for name in [m for m in list(sys.modules) if m == "openpyxl" or m.startswith("openpyxl.")]:
        monkeypatch.delitem(sys.modules, name, raising=False)
    monkeypatch.setitem(sys.modules, "openpyxl", None)
    rc = export_xlsx.main(["--snapshot", str(src), "--out", str(tmp_path / "o.xlsx")])
    assert rc != 0
    assert "pip install openpyxl" in capsys.readouterr().err
