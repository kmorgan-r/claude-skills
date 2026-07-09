#!/usr/bin/env python3
"""Render a snapshot CSV into a formatted .xlsx workbook (Data + Legend sheets).

The snapshot CSV stays the canonical, diff-able source of truth (written by
snapshot.py); this produces a downstream, shareable .xlsx in reports/. openpyxl
is imported lazily so a missing dependency fails loudly (non-zero) instead of
silently, and so the missing-dep path stays testable.

Usage:
    python export_xlsx.py --snapshot data/snapshots/2026-07-03.csv \
                          --out reports/2026-07-03.xlsx
    # --out defaults to reports/<snapshot-basename>.xlsx
"""
import argparse
import csv
import os
import sys

# --- Fill colors (ARGB hex) shared by the Data sheet, the Legend color-key,
#     and the tests. One source of truth so "the key matches the data" holds. ---
FILL_FOUND = "FF63BE7B"       # green  — found / yes / substantive
FILL_TARGET = "FF4F81BD"      # blue   — target
FILL_NOT_FOUND = "FFBFBFBF"   # grey   — not_found
FILL_NO = "FFD65F5F"          # red    — no / symbolic
FILL_YES = FILL_FOUND         # green  — yes / substantive (same green as found)
FILL_HEADER = "FF404040"      # dark   — header row

STATUS_FILLS = {"found": FILL_FOUND, "target": FILL_TARGET, "not_found": FILL_NOT_FOUND}
SMART_YESNO_FILLS = {"yes": FILL_YES, "no": FILL_NO}
SUBSTANCE_FILLS = {"substantive": FILL_YES, "symbolic": FILL_NO}
SMART_COLS = {"smart_specific", "smart_achievable", "smart_relevant"}

# Columns rendered wider, with wrap-text, so long prose stays readable.
WIDE_COLS = {"quote": 60, "assessment_notes": 50, "source_url": 40,
             "linked_targets": 40, "source": 24, "indicator": 24}
WRAP_COLS = {"quote", "assessment_notes"}
DEFAULT_WIDTH = 16
MAX_WIDTH = 60

# --- Legend content (rendered by _build_legend_sheet; tested openpyxl-free) ---
# (column, plain-English meaning, allowed / example values)
LEGEND_COLUMNS = [
    ("entity", "Company / legal entity the row is about.", "Royal Philips"),
    ("lei", "Legal Entity Identifier (GLEIF) — stable ID across renames/mergers.", "H1FJE8H61JGM1JSGM897"),
    ("domain", "ESG domain pack the indicator belongs to.", "climate | circular | biodiversity | social_gov"),
    ("indicator", "Canonical indicator name (see references/indicators.yaml).", "circular_revenue_pct"),
    ("value", "The disclosed number or text value.", "18"),
    ("unit", "Unit for value.", "% | tCO2e | MWh | year | bool"),
    ("period", "Reporting year the value describes; for targets, the target end year.", "2022"),
    ("status", "Whether this is a disclosed actual, a gap, or a forward goal.", "found | not_found | target"),
    ("source", "Human-readable title of the source document.", "Annual Report 2022"),
    ("source_url", "URL of the source document.", "https://…/annual-report-2022.pdf"),
    ("page", "Page / location of the figure in the source.", "p.41"),
    ("quote", "Verbatim snippet from the source containing the number (provenance).", '"circular revenues were 18% of sales"'),
    ("retrieved_at", "Run date the value was pulled (YYYY-MM-DD) — distinct from period.", "2026-06-29"),
    ("item_type", "Kind of item.", "kpi | target | qualitative"),
    ("r_strategy", "Circular R-strategy R0–R9, pipe-separated (primary first). See circular-economy-10rs.json.", "R2|R8"),
    ("enabler_topic", "Enabler that supports circularity (not itself an R-strategy). See circular-economy-10rs.json.", "training | data_infrastructure | traceability | …"),
    ("target_end_year", "Deadline year of a target (YYYY); blank if none stated.", "2025"),
    ("target_has_kpi", "Whether the target carries a quantified KPI (the M of SMART).", "yes | no"),
    ("target_status", "Year-over-year outcome of the target.", "on_track | achieved | delayed | changed | failed | dropped | too_early"),
    ("smart_specific", "SMART — is the target Specific?", "yes | no"),
    ("smart_achievable", "SMART — is the target Achievable?", "yes | no"),
    ("smart_relevant", "SMART — is the target Relevant?", "yes | no"),
    ("substance", "Real operational commitment vs signaling.", "substantive | symbolic"),
    ("planetary_alignment", "Aligned to a planetary boundary / science-based pathway?", "insufficient | pb_aligned | unknown"),
    ("impact_scope", "A–D scoping (NOT GHG Scope 1/2/3).", "A | B | C | D"),
    ("priority_internal", "Strategic priority inside the company.", "high | low"),
    ("importance_external", "External signaling importance.", "high | low"),
    ("linked_targets", "Free text — other targets this connects to, and how.", "netzero_target_year — shares Scope-3 boundary"),
    ("assessment_notes", "Free-text rationale; required whenever any judgment column is set.", "Quantified 25% with a 2025 deadline; board-level KPI."),
]

LEGEND_CODE_TABLES = {
    "status": [
        ("found", "A disclosed actual for that period."),
        ("not_found", "You looked and it wasn't disclosed — a gap is data, record it."),
        ("target", "A forward-looking goal (e.g. \"25% by 2025\")."),
    ],
    "item_type": [
        ("kpi", "A measured metric / indicator."),
        ("target", "A forward-looking goal."),
        ("qualitative", "A narrative commitment with no number."),
    ],
    "impact_scope": [
        ("A", "Footprint — own operations."),
        ("B", "Footprint — direct value chain (suppliers + use phase)."),
        ("C", "Footprint — broader / enabled system."),
        ("D", "Handprint — positive contribution / avoided impact elsewhere."),
    ],
    "planetary_alignment": [
        ("insufficient", "Not aligned to a planetary boundary."),
        ("pb_aligned", "Aligned to a planetary boundary / science-based pathway."),
        ("unknown", "Checked, can't tell (distinct from blank = not assessed)."),
    ],
    "target_status": [
        ("on_track", "Actuals moving toward the target, deadline unchanged."),
        ("achieved", "Target met (an actual now meets/exceeds it)."),
        ("delayed", "Deadline pushed out."),
        ("changed", "Target value or scope restated."),
        ("failed", "Deadline passed without meeting the target."),
        ("dropped", "Target disappeared from disclosure."),
        ("too_early", "First year seen; not yet assessable."),
    ],
}

# Color-key rows for the Legend (fill constant, label). Same FILL_* the Data
# sheet uses, so the swatches match the data cells exactly.
LEGEND_COLOR_KEY = [
    (FILL_FOUND, "found / yes / substantive"),
    (FILL_TARGET, "target"),
    (FILL_NOT_FOUND, "not_found"),
    (FILL_NO, "no / symbolic"),
]

# Pointer note for codes decoded elsewhere (not duplicated into a code table).
LEGEND_POINTERS = [
    ("r_strategy", "R0–R9 circular strategies — see circular-economy-10rs.json."),
    ("enabler_topic", "11 enabler ids — see circular-economy-10rs.json."),
]


def read_snapshot(path):
    """Return (columns_in_file_order, list_of_row_dicts). Reads utf-8-sig so the
    BOM snapshot.py writes is stripped and the first column is 'entity'."""
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        columns = list(reader.fieldnames or [])
        rows = [dict(r) for r in reader]
    return columns, rows


def _cell_fill(name, value):
    """ARGB fill for a data cell, or None (blank, uncolored column, or unknown)."""
    v = (value or "").strip()
    if not v:
        return None
    if name == "status":
        return STATUS_FILLS.get(v)
    if name in SMART_COLS:
        return SMART_YESNO_FILLS.get(v)
    if name == "substance":
        return SUBSTANCE_FILLS.get(v)
    return None


def _solid(color):
    from openpyxl.styles import PatternFill
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _build_data_sheet(ws, columns, rows):
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = _solid(FILL_HEADER)
    for c, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = header_font
        cell.fill = header_fill

    for r, row in enumerate(rows, start=2):
        for c, name in enumerate(columns, start=1):
            value = row.get(name, "")
            cell = ws.cell(row=r, column=c, value=value)
            fill = _cell_fill(name, value)
            if fill:
                cell.fill = _solid(fill)
            if name in WRAP_COLS:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(columns)) if columns else "A"
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

    for c, name in enumerate(columns, start=1):
        width = min(WIDE_COLS.get(name, DEFAULT_WIDTH), MAX_WIDTH)
        ws.column_dimensions[get_column_letter(c)].width = width


def build_workbook(columns, rows):
    """Build a Workbook with a Data sheet (Legend added in Task 3)."""
    from openpyxl import Workbook
    wb = Workbook()
    data = wb.active
    data.title = "Data"
    _build_data_sheet(data, columns, rows)
    return wb
