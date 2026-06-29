# test_linkedin_lead_finder.py
import importlib
import json
import os
import sys
import time

import pytest

sys.path.insert(0, os.path.dirname(__file__))  # import the sibling script

import linkedin_lead_finder as m


# ---------------------------------------------------------------------------
# Shared fakes
# ---------------------------------------------------------------------------

class _Rate:
    def __init__(self, remaining, reset="9999999999"):
        self.remaining = None if remaining is None else str(remaining)
        self.reset = reset


class ApiError(Exception):
    """Stands in for connectsafely.ConnectSafelyError — same message shape."""


class FakeClient:
    """Scriptable get_profile. Each entry: ('ok', profile) | ('api', msg) |
    ('parse', exc). `remaining` parallels the script (None => header missing)."""
    def __init__(self, script, remaining):
        self._script = list(script)
        self._remaining = list(remaining)
        self.calls = []
        self.last_rate = _Rate(100)

    def get_profile(self, profile_id=None, **kw):
        self.calls.append(profile_id)
        kind, payload = self._script.pop(0)
        self.last_rate = _Rate(self._remaining.pop(0))
        if kind == "ok":
            return {"profile": payload}
        if kind == "api":
            raise ApiError(payload)
        if kind == "parse":
            raise payload                     # e.g. ValueError("bad json")


class SourceClient:
    def __init__(self):
        self.calls = []

    def search_people(self, keywords=None, count=25, start=0, **kw):
        self.calls.append(("search_people", start))
        if start == 0:
            return {"people": [{"profileUrl": f"u{i}"} for i in range(count)]}
        return {"people": []}                       # empty 2nd page -> halt

    def search_companies(self, keywords, **kw):
        self.calls.append(("search_companies", keywords))
        return {"companies": [{"companyId": "777"}]}

    def org_followers(self, company_id, **kw):
        self.calls.append(("org_followers", company_id))
        return {"followers": [{"profileUrl": "f1"}]}

    def group_members(self, group_id=None, **kw):
        self.calls.append(("group_members", group_id))
        return {"members": [{"profileUrl": "g1"}]}

    def event_attendees(self, event_id, **kw):
        self.calls.append(("event_attendees", event_id))
        return {}                                   # missing key -> []


class SourceClientForRun:
    last_rate = _Rate(50)

    def search_people(self, keywords=None, count=25, start=0, **kw):
        if start == 0:
            return {"people": [
                {"profileUrl": "https://www.linkedin.com/in/already-have",
                 "firstName": "Dup", "headline": "Sustainability Lead"},
                {"profileUrl": "https://www.linkedin.com/in/fresh-person",
                 "firstName": "Fresh", "headline": "Head of Sustainability"},
            ]}
        return {"people": []}

    def get_profile(self, profile_id=None, **kw):
        self.last_rate = _Rate(40)
        return {"profile": {"currentCompany": "Acme",
                            "aboutText": "bio",
                            "experience": [{"title": "Head of Sustainability"}]}}


# ---------------------------------------------------------------------------
# Task 1: Scaffold module, lazy client, pre-flight
# ---------------------------------------------------------------------------

def test_module_imports_with_key_unset(monkeypatch):
    monkeypatch.delenv("CONNECTSAFELY_API_KEY", raising=False)
    saved = sys.modules.pop("linkedin_lead_finder", None)
    try:
        mod = importlib.import_module("linkedin_lead_finder")
        assert hasattr(mod, "get_client")
        assert hasattr(mod, "preflight")
    finally:
        # restore the canonical module object so later tests share one identity
        if saved is not None:
            sys.modules["linkedin_lead_finder"] = saved


def test_preflight_raises_clear_error_without_key(monkeypatch):
    import linkedin_lead_finder as m
    monkeypatch.delenv("CONNECTSAFELY_API_KEY", raising=False)
    m._CLIENT = None
    with pytest.raises(RuntimeError, match="CONNECTSAFELY_API_KEY"):
        m.preflight()


# ---------------------------------------------------------------------------
# Task 2: Slug normalization
# ---------------------------------------------------------------------------

def test_normalize_slug_happy_and_case_fold():
    assert m.normalize_slug("https://www.linkedin.com/in/John-Doe/") == "john-doe"
    assert m.normalize_slug("https://www.linkedin.com/in/jane.doe?trk=x") == "jane.doe"
    # trailing locale segment stripped
    assert m.normalize_slug("https://www.linkedin.com/in/john-doe/de") == "john-doe"


def test_normalize_slug_rejects_non_person_and_garbage():
    assert m.normalize_slug("https://www.linkedin.com/company/acme") is None
    assert m.normalize_slug("linkedin.com/company/acme") is None  # scheme-less, no /in/
    assert m.normalize_slug("https://www.linkedin.com/school/mit") is None
    assert m.normalize_slug("https://www.linkedin.com/in/józef") is None  # non-ASCII dropped
    assert m.normalize_slug("") is None
    assert m.normalize_slug(None) is None


def test_secondary_key_canonicalizes():
    assert m.secondary_key("HTTPS://www.LinkedIn.com/in/john-doe/") == \
        "linkedin.com/in/john-doe"
    assert m.secondary_key("") == ""


# ---------------------------------------------------------------------------
# Task 3: Exclude-set build + global dedup
# ---------------------------------------------------------------------------

def test_build_exclude_sets_normalizes_and_counts_drops():
    raw = [
        "https://www.linkedin.com/in/John-Doe/",   # -> slug john-doe
        "https://www.linkedin.com/company/acme",   # invalid slug -> secondary key only
        "",                                        # empty -> dropped from both
    ]
    slugs, secondary, dropped = m.build_exclude_sets(raw)
    assert "john-doe" in slugs
    assert "linkedin.com/company/acme" in secondary
    assert dropped == 1  # the empty string


def test_dedup_suppresses_by_slug_case_insensitive_and_within_batch():
    people = [
        {"profileUrl": "https://www.linkedin.com/in/john-doe"},   # in exclude (case)
        {"profileUrl": "https://www.linkedin.com/in/Alice"},      # fresh
        {"profileUrl": "https://www.linkedin.com/in/alice/"},     # within-batch dup
    ]
    slugs = {"john-doe"}
    survivors, dropped = m.dedup(people, slugs, set())
    assert [s["slug"] for s in survivors] == ["alice"]
    assert len(dropped) == 2


def test_dedup_secondary_key_suppresses_malformed_existing():
    # An existing Odoo contact whose stored URL failed slug validation but is
    # real; a freshly-sourced version of the SAME person must be dropped and
    # never reach the survivor (enrich) set.
    people = [{"profileUrl": "https://www.linkedin.com/company/acme-person"}]
    survivors, dropped = m.dedup(
        people, set(), {"linkedin.com/company/acme-person"}
    )
    assert survivors == []
    assert len(dropped) == 1


# ---------------------------------------------------------------------------
# Task 4: Cheap ICP scorer + threshold partition
# ---------------------------------------------------------------------------

def test_cheap_score_counts_distinct_keyword_hits():
    p = {"headline": "Head of Sustainability", "currentPosition": "Sustainability Lead",
         "location": "Berlin"}
    assert m.cheap_score(p, ["sustainability", "carbon"]) == 1  # distinct keyword


def test_cheap_score_tolerates_missing_title_key():
    p = {"headline": "Carbon Accounting Manager"}  # no currentPosition
    assert m.cheap_score(p, ["carbon"]) == 1


def test_cheap_score_word_boundary_no_substring_false_positives():
    p = {"headline": "Smart Cities Trainee", "location": "Stuttgart"}
    # "ai" must NOT match "Trainee"; "art" must NOT match "Smart"/"Stuttgart"
    assert m.cheap_score(p, ["ai", "art"]) == 0


def test_cheap_score_blank_keyword_matches_nothing():
    # an empty/whitespace keyword must NOT pass every lead (would be r"\b\b")
    assert m.cheap_score({"headline": "Software Engineer"}, ["", "  "]) == 0


def test_score_and_partition_threshold_zero_floored_to_one():
    people = [{"headline": "Software Engineer"}]   # score 0 vs "sustainability"
    survivors, rejected = m.score_and_partition(people, ["sustainability"], threshold=0)
    assert survivors == [] and len(rejected) == 1  # threshold floored to 1, score 0 rejected


def test_score_and_partition_threshold_and_zero():
    people = [
        {"headline": "Sustainability Director"},     # score 1
        {"headline": "Software Engineer"},           # score 0 -> rejected
    ]
    survivors, rejected = m.score_and_partition(people, ["sustainability"], threshold=1)
    assert len(survivors) == 1 and survivors[0]["cheap_score"] == 1
    assert len(rejected) == 1  # score 0 excluded from enrich set


# ---------------------------------------------------------------------------
# Task 5: Profile field extraction
# ---------------------------------------------------------------------------

def test_extract_profile_fields_uses_real_keys():
    profile = {
        "currentCompany": "Acme GmbH",
        "aboutText": "Sustainability leader.\nDriving PCF.",
        "topSkills": ["LCA", "ISO 14067"],
        "experience": [{"title": "Head of Sustainability", "companyName": "Acme GmbH"}],
    }
    out = m.extract_profile_fields(profile)
    assert out["company_name"] == "Acme GmbH"
    assert out["x_summary"] == "Sustainability leader. Driving PCF."  # newlines flattened
    assert out["x_job_title"] == "Head of Sustainability"
    assert out["top_skills"] == "LCA, ISO 14067"


def test_extract_profile_fields_nulls_and_missing_keys_blank():
    out = m.extract_profile_fields({})  # missing everything
    assert out == {"company_name": "", "x_summary": "", "x_job_title": "", "top_skills": ""}
    # A wrong-key fixture (legacy 'about') must NOT populate x_summary.
    assert m.extract_profile_fields({"about": "x"})["x_summary"] == ""


# ---------------------------------------------------------------------------
# Task 6: Enrich checkpoint
# ---------------------------------------------------------------------------

def test_load_checkpoint_missing_returns_initial(tmp_path):
    st = m.load_checkpoint(str(tmp_path / "none.json"))
    assert st == {"done": set(), "count": 0, "reset": None}


def test_save_is_atomic_and_roundtrips(tmp_path):
    p = str(tmp_path / "ckpt.json")
    m.save_checkpoint(p, {"done": {"a", "b"}, "count": 2, "reset": "1000"})
    st = m.load_checkpoint(p)
    assert st["done"] == {"a", "b"} and st["count"] == 2 and st["reset"] == "1000"
    # no leftover temp file
    assert not any(str(f).endswith(".tmp") for f in tmp_path.iterdir())


def test_counter_for_run_resets_after_reset_passes():
    # reset in the past -> counter zeroes (fresh budget)
    assert m.counter_for_run({"count": 90, "reset": "1000"}, now=2000.0) == 0
    # reset in the future -> keep persisted count (cap still enforced)
    assert m.counter_for_run({"count": 90, "reset": "9999999999"}, now=2000.0) == 90
    # no reset persisted -> treat as fresh
    assert m.counter_for_run({"count": 5, "reset": None}, now=2000.0) == 0


# ---------------------------------------------------------------------------
# Task 7: Enrich loop
# ---------------------------------------------------------------------------

def test_enrich_only_survivors_and_resumes_skipping_done(tmp_path):
    ckpt = str(tmp_path / "c.json")
    m.save_checkpoint(ckpt, {"done": {"a"}, "count": 1, "reset": "9999999999"})
    survivors = [{"slug": "a"}, {"slug": "b"}]
    client = FakeClient([("ok", {"currentCompany": "X"})], remaining=[50])
    out, state = m.enrich(survivors, client, ckpt, cap=120, floor=5,
                          now_fn=lambda: 0.0)
    assert client.calls == ["b"]                  # 'a' already done, skipped (no call)
    # 'a' is flagged enriched (resumed from a prior run) AND 'b' is freshly enriched
    assert {s["slug"] for s in out if s.get("enriched")} == {"a", "b"}
    assert next(s for s in out if s["slug"] == "a")["resumed"] is True
    assert "b" in state["done"]


def test_enrich_cap_429_stops_day(tmp_path):
    ckpt = str(tmp_path / "c.json")
    survivors = [{"slug": "a"}, {"slug": "b"}]
    client = FakeClient([("api", "POST /profile -> 429: rate limit")],
                        remaining=[4])
    out, state = m.enrich(survivors, client, ckpt, now_fn=lambda: 0.0)
    assert client.calls == ["a"]                  # stopped after the 429
    assert "a" not in state["done"]               # cap failure not marked done
    # the freshest server reset is persisted for the next run's window math
    reloaded = m.load_checkpoint(ckpt)
    assert reloaded["reset"] == "9999999999"


def test_enrich_transient_retries_once_then_skips(tmp_path):
    ckpt = str(tmp_path / "c.json")
    survivors = [{"slug": "a"}]
    client = FakeClient(
        [("api", "POST /profile -> 503: busy"), ("api", "POST /profile -> 503: busy")],
        remaining=[50, 50])
    out, state = m.enrich(survivors, client, ckpt, now_fn=lambda: 0.0)
    assert client.calls == ["a", "a"]             # one retry, then skip
    assert "a" not in state["done"]


def test_enrich_parse_error_skips_without_retry(tmp_path):
    ckpt = str(tmp_path / "c.json")
    survivors = [{"slug": "a"}, {"slug": "b"}]
    client = FakeClient(
        [("parse", ValueError("Expecting value")), ("ok", {"currentCompany": "X"})],
        remaining=[50, 50])
    out, state = m.enrich(survivors, client, ckpt, now_fn=lambda: 0.0)
    assert client.calls == ["a", "b"]             # 'a' parse-failed, NO retry; 'b' ran
    assert "a" not in state["done"] and "b" in state["done"]
    assert any(r.get("slug") == "a" and r.get("enrich_error") for r in out)


def test_enrich_non_dict_profile_does_not_abort_batch(tmp_path):
    # A 2xx body that is valid JSON but the wrong shape (profile is an int) makes
    # extract_profile_fields do .get on a non-dict -> AttributeError -> must be
    # caught as a parse error, not crash the batch.
    ckpt = str(tmp_path / "c.json")
    survivors = [{"slug": "a"}, {"slug": "b"}]
    client = FakeClient([("ok-raw", 123), ("ok", {"currentCompany": "X"})],
                        remaining=[50, 50])
    # 'ok-raw' returns {"profile": 123}
    client._script[0] = ("ok", 123)               # profile payload is an int
    out, state = m.enrich(survivors, client, ckpt, now_fn=lambda: 0.0)
    assert client.calls == ["a", "b"]
    assert "b" in state["done"]                   # batch continued past the bad row


def test_enrich_live_floor_hard_stops(tmp_path):
    ckpt = str(tmp_path / "c.json")
    survivors = [{"slug": "a"}, {"slug": "b"}]
    client = FakeClient([("ok", {"currentCompany": "X"})], remaining=[5])
    out, state = m.enrich(survivors, client, ckpt, floor=5, now_fn=lambda: 0.0)
    assert client.calls == ["a"]                  # floor hit after first call


def test_enrich_no_deadlock_with_future_reset_at_cap(tmp_path):
    # The deadlock guard: persisted count >= cap with a FUTURE reset must NOT
    # block the first call — the live floor is the only hard stop.
    ckpt = str(tmp_path / "c.json")
    m.save_checkpoint(ckpt, {"done": set(), "count": 120, "reset": "9999999999"})
    survivors = [{"slug": "a"}]
    client = FakeClient([("ok", {"currentCompany": "X"})], remaining=[50])
    out, state = m.enrich(survivors, client, ckpt, cap=120, floor=5,
                          now_fn=lambda: 0.0)
    assert client.calls == ["a"]                  # probed despite count>=cap


def test_enrich_missing_header_falls_back_to_cap(tmp_path):
    # When the rate header is absent (remaining None), the live floor can't fire;
    # the local cap must serve as a fallback hard stop (after >=1 call).
    ckpt = str(tmp_path / "c.json")
    survivors = [{"slug": "a"}, {"slug": "b"}, {"slug": "c"}]
    client = FakeClient(
        [("ok", {"currentCompany": "X"}), ("ok", {"currentCompany": "Y"})],
        remaining=[None, None])
    out, state = m.enrich(survivors, client, ckpt, cap=2, floor=5, now_fn=lambda: 0.0)
    assert client.calls == ["a", "b"]             # stopped at cap=2, no infinite run


# ---------------------------------------------------------------------------
# Task 8: Selection-key coercion + seniority derivation
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("key", sorted({
    "sustainability", "product_rd", "ops_sc", "founder_exec", "investor",
    "marketing", "technical", "partner", "low_fit", "unknown"}))
def test_all_personas_pass_through(key):
    assert m.coerce_persona(key) == key


def test_invalid_persona_coerces_unknown():
    assert m.coerce_persona("c-suite") == "unknown"
    assert m.coerce_persona(None) == "unknown"


@pytest.mark.parametrize("key", ["analyst", "manager", "director", "vp", "c_level"])
def test_all_seniorities_pass_through(key):
    assert m.validate_seniority(key) == key


def test_invalid_seniority_returns_none():
    assert m.validate_seniority("exec") is None       # near-miss, not a real key
    assert m.validate_seniority(None) is None


def test_derive_seniority_maps_and_defaults_none():
    assert m.derive_seniority("Chief Sustainability Officer") == "c_level"
    assert m.derive_seniority("VP of Engineering") == "vp"
    assert m.derive_seniority("Head of Sustainability") == "director"
    assert m.derive_seniority("Sustainability Manager") == "manager"
    assert m.derive_seniority("Sustainability Analyst") == "analyst"
    assert m.derive_seniority("Consultant") is None   # unmappable -> omit


def test_derive_seniority_word_boundary_no_substring_collisions():
    assert m.derive_seniority("VPN Security Engineer") is None  # "vp" not in "VPN"
    assert m.derive_seniority("Team Lead") == "manager"         # \blead\b matches


# ---------------------------------------------------------------------------
# Task 9: Cell neutralization + create-payload builder + schema verify
# ---------------------------------------------------------------------------

def test_neutralize_prefixes_formula_starters():
    assert m.neutralize("=SUM(A1)") == "'=SUM(A1)"
    assert m.neutralize("@cmd") == "'@cmd"
    assert m.neutralize("\tx") == "'\tx"
    assert m.neutralize("safe") == "safe"
    assert m.neutralize(5) == "5"


def test_build_payload_coerces_persona_and_omits_seniority_and_country():
    lead = {"slug": "jane", "first_name": "Jane", "last_name": "Doe",
            "x_headline": "Head of Sustainability", "company_name": "Acme",
            "x_summary": "bio", "x_persona": "bogus", "x_seniority": "exec",
            "x_need_state": "PCF", "x_lead_score": 8, "x_outreach_angle": "angle",
            "profileUrl": "https://www.linkedin.com/in/jane"}
    payload = m.build_payload(lead, country_id=None)
    assert payload["x_persona"] == "unknown"          # invalid coerced
    assert "x_seniority" not in payload                # invalid omitted
    assert "country_id" not in payload                 # no-match omitted
    assert payload["x_lead_status"] == "New"
    assert payload["email"] == ""
    assert payload["x_linkedin_url"] == lead["profileUrl"]


def test_build_payload_includes_country_when_resolved():
    lead = {"x_persona": "sustainability", "x_seniority": "director"}
    payload = m.build_payload(lead, country_id=42)
    assert payload["country_id"] == 42
    assert payload["x_seniority"] == "director"


def test_verify_schema_raises_on_missing_field():
    present = set(m.REQUIRED_FIELDS) - {"x_summary"}
    with pytest.raises(RuntimeError, match="x_summary"):
        m.verify_schema(present)
    m.verify_schema(set(m.REQUIRED_FIELDS))            # all present -> no raise


# ---------------------------------------------------------------------------
# Task 10: Workbook writer
# ---------------------------------------------------------------------------

def test_write_workbook_sheets_columns_and_neutralized(tmp_path):
    import openpyxl
    path = str(tmp_path / "out.xlsx")
    leads = [{"first_name": "Jane", "x_summary": "=DANGER()", "odoo_ready": "no"}]
    rejected = [{"reject_reason": "@evil from API body", "slug": "x"}]
    m.write_workbook(path, leads, rejected, {"mode": "people", "keywords": "carbon"})
    wb = openpyxl.load_workbook(path)
    assert set(wb.sheetnames) == {"Leads", "Rejected", "Run Config"}
    leads_ws = wb["Leads"]
    header = [c.value for c in leads_ws[1]]
    assert "odoo_ready" in header and "x_summary" in header
    # the formula-leading cell is neutralized
    col = header.index("x_summary")
    assert leads_ws.cell(row=2, column=col + 1).value == "'=DANGER()"
    # untrusted reject reason (API body) neutralized on the Rejected sheet too
    rej_ws = wb["Rejected"]
    rej_header = [c.value for c in rej_ws[1]]
    rcol = rej_header.index("reject_reason")
    assert rej_ws.cell(row=2, column=rcol + 1).value == "'@evil from API body"


# ---------------------------------------------------------------------------
# Task 11: Source dispatch + pagination
# ---------------------------------------------------------------------------

def test_source_people_paginates_and_halts():
    c = SourceClient()
    out = m.source_people(c, "carbon", page_size=25)
    assert len(out) == 25
    assert ("search_people", 0) in c.calls and ("search_people", 25) in c.calls


def test_dispatch_org_followers_resolves_company_first():
    c = SourceClient()
    out = m.dispatch_source("org-followers", c, keywords="Acme")
    assert ("search_companies", "Acme") in c.calls
    assert ("org_followers", "777") in c.calls
    assert out == [{"profileUrl": "f1"}]


def test_dispatch_event_missing_key_returns_empty():
    c = SourceClient()
    assert m.dispatch_source("event", c, event_id="e1") == []
    assert ("event_attendees", "e1") in c.calls


def test_source_event_reads_attendees_key():
    class C:
        def event_attendees(self, event_id, **kw):
            return {"attendees": [{"profileUrl": "a1"}]}
    assert m.source_event(C(), "e1") == [{"profileUrl": "a1"}]


# ---------------------------------------------------------------------------
# Task 12: CLI wiring + end-to-end dry run
# ---------------------------------------------------------------------------

def test_parse_args_defaults():
    a = m.parse_args(["--mode", "people", "--keywords", "carbon",
                      "--out", "x.xlsx"])
    assert a.mode == "people" and a.keywords == "carbon"
    assert a.threshold == 1 and a.cap == 120 and a.floor == 5


def test_run_end_to_end_dry(tmp_path):
    out = str(tmp_path / "leads.xlsx")
    exclude = str(tmp_path / "exclude.txt")
    with open(exclude, "w", encoding="utf-8") as f:
        f.write("https://www.linkedin.com/in/already-have\n")
    args = m.parse_args(["--mode", "people", "--keywords", "sustainability",
                         "--exclude-file", exclude, "--out", out,
                         "--checkpoint", str(tmp_path / "c.json"),
                         "--keyword-score", "sustainability"])
    client = SourceClientForRun()       # see below
    path = m.run(args, client=client)
    assert os.path.exists(path)
    import openpyxl
    wb = openpyxl.load_workbook(path)
    assert "Leads" in wb.sheetnames
    ws = wb["Leads"]
    header = [c.value for c in ws[1]]
    rows = [{header[i]: c.value for i, c in enumerate(r)} for r in ws.iter_rows(min_row=2)]
    # the excluded dup ("already-have") is gone; only the fresh person survives, enriched
    assert len(rows) == 1
    assert rows[0]["slug"] == "fresh-person"
    assert rows[0]["company_name"] == "Acme"     # enrich populated it
    assert str(rows[0]["enriched"]) in ("True", "1")


def test_run_neutralizes_malicious_profile_content_end_to_end(tmp_path):
    # The realistic injection vector: attacker-controlled profile content (currentCompany
    # / aboutText) that BEGINS with a formula char must land neutralized on the Leads
    # sheet. (enrich_error carries a human-readable prefix and is inert by construction,
    # so it's the wrong column to assert on — profile content is the live vector.)
    out = str(tmp_path / "leads.xlsx")
    args = m.parse_args(["--mode", "people", "--keywords", "sustainability",
                         "--out", out, "--checkpoint", str(tmp_path / "c.json"),
                         "--keyword-score", "sustainability"])

    class MaliciousProfileClient:
        last_rate = _Rate(50)
        def search_people(self, keywords=None, count=25, start=0, **kw):
            if start == 0:
                return {"people": [{"profileUrl": "https://www.linkedin.com/in/x",
                                    "headline": "Head of Sustainability"}]}
            return {"people": []}
        def get_profile(self, profile_id=None, **kw):
            self.last_rate = _Rate(50)
            return {"profile": {"currentCompany": "=cmd|'/c calc'!A1",
                                "aboutText": "@SUM(1)"}}

    m.run(args, client=MaliciousProfileClient())
    import openpyxl
    ws = openpyxl.load_workbook(out)["Leads"]
    header = [c.value for c in ws[1]]
    crow = [c.value for c in ws[2]]
    assert crow[header.index("company_name")] == "'=cmd|'/c calc'!A1"   # neutralized
    assert crow[header.index("x_summary")] == "'@SUM(1)"                 # neutralized


def test_run_schema_manifest_fails_fast(tmp_path):
    import json as _json
    manifest = str(tmp_path / "schema.json")
    with open(manifest, "w", encoding="utf-8") as f:
        _json.dump(sorted(set(m.REQUIRED_FIELDS) - {"x_summary"}), f)
    args = m.parse_args(["--mode", "people", "--keywords", "x", "--out",
                         str(tmp_path / "o.xlsx"), "--schema-manifest", manifest,
                         "--checkpoint", str(tmp_path / "c.json")])
    with pytest.raises(RuntimeError, match="x_summary"):
        m.run(args, client=SourceClientForRun())
