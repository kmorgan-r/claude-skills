"""LinkedIn-native B2B lead sourcer → Odoo-ready mailing.contact rows.

Pure pipeline: source (ConnectSafely) → dedup → cheap ICP filter → capped enrich
→ classify prep → workbook. SAFE BY DEFAULT: no sends, no email, no web scrape.
All Odoo MCP work is agent-side (see SKILL.md).
"""
import argparse
import json
import os
import re
import sys
import time

import openpyxl

# connectsafely.py lives in the marketing dir, not this repo.
MARKETING_DIR = os.environ.get(
    "MARKETING_DIR", os.path.expanduser(r"~\marketing")
)

_CLIENT = None

_SLUG_RE = re.compile(r"^[A-Za-z0-9._-]+$")
_NON_PERSON = ("/company/", "/school/", "/showcase/")

# Parse-failure exception types: a malformed/wrong-shape 2xx body. extract_profile_fields
# does .get() on the profile, so a non-dict profile raises AttributeError — include it.
_PARSE_ERRORS = (ValueError, KeyError, TypeError, AttributeError)

VALID_PERSONAS = frozenset({
    "sustainability", "product_rd", "ops_sc", "founder_exec", "investor",
    "marketing", "technical", "partner", "low_fit", "unknown",
})
VALID_SENIORITIES = frozenset({"analyst", "manager", "director", "vp", "c_level"})

_FORMULA_STARTERS = ("=", "+", "-", "@", "\t", "\r", "\n")

REQUIRED_FIELDS = frozenset({
    "x_linkedin_url", "first_name", "last_name", "x_headline", "x_job_title",
    "company_name", "x_summary", "x_persona", "x_need_state", "x_outreach_angle",
    "x_lead_score", "x_lead_status",
})

LEAD_COLUMNS = [
    "slug", "profileUrl", "first_name", "last_name", "x_headline", "x_job_title",
    "company_name", "x_summary", "x_seniority", "seniority_unset", "x_persona",
    "x_need_state", "x_lead_score", "x_outreach_angle", "x_industry",
    "x_department_function", "location", "cheap_score", "enriched", "enrich_error",
    "x_lead_status", "email", "odoo_ready", "created",
]
REJECT_COLUMNS = ["slug", "profileUrl", "first_name", "last_name", "x_headline",
                  "cheap_score", "reject_reason"]

MAX_SOURCE_RESULTS = 200  # bound every mode so a huge org/group/event can't blow the
                          # enrich budget in one run (the live floor may never descend
                          # to `floor` within an oversized survivor list).


# ---------------------------------------------------------------------------
# Client
# ---------------------------------------------------------------------------

def get_client():
    """Lazily construct + cache the ConnectSafely client.

    The client module instantiates at import and sys.exit()s if the key is
    unset; we only import it here, never at module top, so the script and its
    tests import cleanly without a key.
    """
    global _CLIENT
    if _CLIENT is None:
        if MARKETING_DIR not in sys.path:
            sys.path.insert(0, MARKETING_DIR)
        from connectsafely import cs  # constructed at import of that module
        _CLIENT = cs
    return _CLIENT


def preflight():
    """Fail fast with a clear message if the API key is absent.

    Eagerly constructs the client once inside a SystemExit guard, converting the
    client's import-time sys.exit() into a catchable RuntimeError. After this the
    client is cached, so later get_client() calls never re-enter the constructor.
    """
    if not os.environ.get("CONNECTSAFELY_API_KEY"):
        raise RuntimeError(
            "CONNECTSAFELY_API_KEY not set — set it in the environment before running."
        )
    try:
        get_client()
    except SystemExit as e:
        raise RuntimeError(
            "CONNECTSAFELY_API_KEY not set — set it in the environment before running."
        ) from e


# ---------------------------------------------------------------------------
# Slug normalization
# ---------------------------------------------------------------------------

def normalize_slug(url):
    if not url or "/in/" not in url:
        return None
    if any(seg in url for seg in _NON_PERSON):
        return None
    slug = url.split("/in/")[-1].rstrip("/").split("?")[0].split("/")[0].lower()
    if not slug or not _SLUG_RE.match(slug):
        return None
    return slug


def secondary_key(url):
    if not url:
        return ""
    s = url.strip().lower()
    s = re.sub(r"^https?://", "", s)
    s = re.sub(r"^www\.", "", s)
    return s.split("?")[0].rstrip("/")


# ---------------------------------------------------------------------------
# Exclude-set build + global dedup
# ---------------------------------------------------------------------------

def build_exclude_sets(raw_urls):
    slugs, secondary, dropped = set(), set(), 0
    for url in raw_urls:
        slug = normalize_slug(url)
        if slug:
            slugs.add(slug)
            continue
        key = secondary_key(url)
        if key:
            secondary.add(key)
        else:
            dropped += 1
    return slugs, secondary, dropped


def dedup(people, exclude_slugs, exclude_secondary):
    survivors, dropped = [], []
    seen_slugs, seen_secondary = set(), set()
    for p in people:
        url = p.get("profileUrl", "")
        slug = normalize_slug(url)
        if slug:
            if slug in exclude_slugs or slug in seen_slugs:
                dropped.append(p)
                continue
            seen_slugs.add(slug)
            survivors.append({**p, "slug": slug})
        else:
            key = secondary_key(url)
            if not key or key in exclude_secondary or key in seen_secondary:
                dropped.append(p)
                continue
            seen_secondary.add(key)
            # mark secondary-key survivors: their "slug" is not a real profile id,
            # so the enrich stage must skip get_profile for them.
            survivors.append({**p, "slug": key, "secondary": True})
    return survivors, dropped


# ---------------------------------------------------------------------------
# Cheap ICP scorer + threshold partition
# ---------------------------------------------------------------------------

def cheap_score(person, keywords):
    # Word-boundary match, NOT substring: a substring match would count "ai" inside
    # "email" or "art" inside "Stuttgart", polluting the scarce enrich set with
    # non-ICP people. \b anchors each keyword (phrases like "head of sustainability"
    # still match as a unit).
    parts = [
        person.get("headline", "") or "",
        person.get("currentPosition", "") or "",
        person.get("location", "") or "",
    ]
    hay = " ".join(parts).lower()
    # Drop blank/whitespace keywords: an empty keyword compiles to r"\b\b", which
    # matches any haystack and would pass EVERY lead through the ICP filter.
    kws = [k for k in (kw.strip().lower() for kw in keywords) if k]
    return sum(1 for kw in kws if re.search(r"\b" + re.escape(kw) + r"\b", hay))


def score_and_partition(people, keywords, threshold):
    # A lead is kept iff score >= threshold AND has at least one ICP hit (score > 0).
    # threshold is floored at 1 (an ICP filter with no required signal is meaningless);
    # threshold=0 is treated as 1 so "score 0 -> rejected" always holds.
    threshold = max(int(threshold), 1)
    survivors, rejected = [], []
    for p in people:
        score = cheap_score(p, keywords)
        if score >= threshold:
            survivors.append({**p, "cheap_score": score})
        else:
            rejected.append({**p, "cheap_score": score})
    return survivors, rejected


# ---------------------------------------------------------------------------
# Profile field extraction (real get_profile shape)
# ---------------------------------------------------------------------------

def extract_profile_fields(profile):
    profile = profile or {}
    exp = profile.get("experience") or []
    title = exp[0].get("title", "") if exp and isinstance(exp[0], dict) else ""
    return {
        "company_name": profile.get("currentCompany") or "",
        "x_summary": (profile.get("aboutText") or "").replace("\n", " ").strip(),
        "x_job_title": title or "",
        "top_skills": ", ".join(profile.get("topSkills") or []),
    }


# ---------------------------------------------------------------------------
# Enrich checkpoint (atomic load/save + reset logic)
# ---------------------------------------------------------------------------

def load_checkpoint(path):
    if not os.path.exists(path):
        return {"done": set(), "count": 0, "reset": None}
    with open(path, encoding="utf-8") as f:
        raw = json.load(f)
    return {
        "done": set(raw.get("done", [])),
        "count": int(raw.get("count", 0)),
        "reset": raw.get("reset"),
    }


def save_checkpoint(path, state):
    tmp = path + ".tmp"
    payload = {
        "done": sorted(state.get("done", [])),
        "count": int(state.get("count", 0)),
        "reset": state.get("reset"),
    }
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(payload, f)
    os.replace(tmp, path)


def counter_for_run(state, now):
    reset = state.get("reset")
    if not reset:
        return 0
    try:
        if now >= float(reset):
            return 0
    except (TypeError, ValueError):
        return 0
    return int(state.get("count", 0))


# ---------------------------------------------------------------------------
# Helpers used by enrich
# ---------------------------------------------------------------------------

def _safe_int(rate):
    if rate is None:
        return None
    try:
        return int(rate.remaining)
    except (TypeError, ValueError, AttributeError):
        return None


def _is_cap(msg):
    return "429" in msg or "rate limit" in msg.lower()


# ---------------------------------------------------------------------------
# Enrich loop (budget cap, cap/transient/parse discrimination, live floor)
# ---------------------------------------------------------------------------

def enrich(survivors, client, checkpoint_path, *, cap=120, floor=5, now_fn=time.time):
    """Enrich post-dedup, post-filter survivors with get_profile, budget-capped.

    Does NOT import connectsafely: parse failures are caught by exception TYPE
    (_PARSE_ERRORS); API errors are any other Exception, classified cap-vs-transient
    by message via _is_cap. The post-call live floor (cs.last_rate.remaining) is the
    sole authoritative hard stop; the local `count` is a fallback only for when the
    header is missing, and is never checked BEFORE a call (so it can never deadlock).
    """
    state = load_checkpoint(checkpoint_path)
    count = counter_for_run(state, now_fn())
    out = []
    for i, s in enumerate(survivors):
        slug = s["slug"]
        if slug in state["done"]:
            # already enriched on a PRIOR run — the extracted fields are NOT in this
            # workbook (the checkpoint stores only slugs). Mark it resumed so run()
            # forces odoo_ready=no; a human must never create a blank mailing.contact
            # from a field-empty resumed row.
            out.append({**s, "enriched": True, "resumed": True,
                        "enrich_error": "enriched on a prior run — fields not in this workbook"})
            continue
        if s.get("secondary"):        # malformed-URL dup key — not a real profile slug
            out.append({**s, "enriched": False, "enrich_error": "no valid profile slug"})
            continue
        profile, err = None, None
        for attempt in (1, 2):
            spent = False             # did this attempt's call reach a 2xx?
            try:
                resp = client.get_profile(profile_id=slug)
                spent = True
                count += 1            # a 2xx call spends exactly one quota unit
                profile = extract_profile_fields((resp or {}).get("profile") or {})
                break
            except _PARSE_ERRORS as e:
                if not spent:         # parse failure inside get_profile's own json()
                    count += 1
                err = f"parse error: {e}"
                break                 # deterministic — no retry, no double-count
            except Exception as e:    # API error (e.g. ConnectSafelyError)
                count += 1
                if _is_cap(str(e)):
                    state["count"] = count
                    state["reset"] = getattr(getattr(client, "last_rate", None),
                                             "reset", state["reset"])
                    save_checkpoint(checkpoint_path, state)
                    rest = [{**r, "enriched": False,
                             "enrich_error": "skipped: daily enrich cap reached"}
                            for r in survivors[i:]]
                    rest[0]["enrich_error"] = f"cap reached: {e}"
                    return out + rest, state
                err = f"api error: {e}"
                if attempt == 2:      # transient, retry exhausted
                    break
        remaining = _safe_int(getattr(client, "last_rate", None))
        if profile is not None:
            out.append({**s, **profile, "enriched": True})
            state["done"].add(slug)
            state["count"] = count
            state["reset"] = getattr(client.last_rate, "reset", state["reset"])
            save_checkpoint(checkpoint_path, state)
        else:
            out.append({**s, "enriched": False, "enrich_error": err or "unknown"})
        # Hard stops, BOTH post-call (never a pre-call block, so no deadlock):
        if remaining is not None and remaining <= floor:
            break                     # authoritative live floor (the common case)
        if count >= cap:
            break                     # hard ceiling — bounds the uncapped source modes
    return out, state


# ---------------------------------------------------------------------------
# Selection-key coercion + seniority derivation
# ---------------------------------------------------------------------------

def coerce_persona(key):
    return key if key in VALID_PERSONAS else "unknown"


def validate_seniority(key):
    return key if key in VALID_SENIORITIES else None


def _has_word(t, *words):
    return any(re.search(r"\b" + w + r"\b", t) for w in words)


def derive_seniority(title):
    # Word-boundary matching for short/ambiguous tokens: "vp" must not match "VPN",
    # "lead" must not match a longer word. Order matters (most senior first).
    t = (title or "").lower()
    if _has_word(t, "chief", "cxo", "ceo", "cto", "cfo", "coo", "cmo") or "c-level" in t:
        return "c_level"
    if _has_word(t, "vp") or "vice president" in t:
        return "vp"
    if "head of" in t or _has_word(t, "director"):
        return "director"
    if _has_word(t, "manager", "lead"):
        return "manager"
    if _has_word(t, "analyst", "associate"):
        return "analyst"
    return None


# ---------------------------------------------------------------------------
# Cell neutralization + create-payload builder + schema verify
# ---------------------------------------------------------------------------

def neutralize(value):
    s = "" if value is None else str(value)
    if s and s[0] in _FORMULA_STARTERS:
        return "'" + s
    return s


def build_payload(lead, country_id=None):
    payload = {
        "x_linkedin_url": lead.get("profileUrl", ""),
        "first_name": lead.get("first_name", ""),
        "last_name": lead.get("last_name", ""),
        "x_headline": lead.get("x_headline", ""),
        "x_job_title": lead.get("x_job_title", ""),
        "company_name": lead.get("company_name", ""),
        "x_summary": lead.get("x_summary", ""),
        "x_persona": coerce_persona(lead.get("x_persona")),
        "x_need_state": lead.get("x_need_state", ""),
        "x_outreach_angle": lead.get("x_outreach_angle", ""),
        "x_lead_score": lead.get("x_lead_score", 0),
        "x_lead_status": "New",
        "email": "",
    }
    seniority = validate_seniority(lead.get("x_seniority"))
    if seniority is not None:
        payload["x_seniority"] = seniority
    if country_id is not None:
        payload["country_id"] = country_id
    return payload


def verify_schema(present):
    missing = sorted(set(REQUIRED_FIELDS) - set(present))
    if missing:
        raise RuntimeError(
            "mailing.contact missing required field(s): " + ", ".join(missing)
        )


# ---------------------------------------------------------------------------
# Workbook writer (Leads/Rejected/Run Config, all cells neutralized)
# ---------------------------------------------------------------------------

def _write_sheet(ws, columns, rows):
    ws.append(columns)
    for row in rows:
        ws.append([neutralize(row.get(c, "")) for c in columns])


def write_workbook(path, leads, rejected, run_config):
    wb = openpyxl.Workbook()
    leads_ws = wb.active
    leads_ws.title = "Leads"
    _write_sheet(leads_ws, LEAD_COLUMNS, leads)
    _write_sheet(wb.create_sheet("Rejected"), REJECT_COLUMNS, rejected)
    cfg = wb.create_sheet("Run Config")
    cfg.append(["key", "value"])
    for k, v in run_config.items():
        cfg.append([neutralize(k), neutralize(v)])
    wb.save(path)


# ---------------------------------------------------------------------------
# Source dispatch + pagination
# ---------------------------------------------------------------------------

def source_people(client, keywords, filters=None, page_size=25, max_results=200):
    people, start = [], 0
    while len(people) < max_results:
        resp = client.search_people(keywords=keywords, count=page_size,
                                    start=start, filters=filters)
        batch = (resp or {}).get("people") or []
        if not batch:
            break
        people.extend(batch)
        if len(batch) < page_size:
            break
        start += page_size
    return people[:max_results]


# CONFIRM AT IMPLEMENTATION: the search_companies wrapper key ("companies") and the
# id field ("companyId"/"id") are not pinned in connectsafely.py — verify both against a
# real response (or a captured fixture) before relying on them. Likewise confirm
# cs.last_rate exposes .remaining/.reset (the enrich hard stop depends on it). The guards
# below fail LOUD on an unexpected shape rather than coercing None -> "None".
def resolve_company_id(client, keywords):
    resp = client.search_companies(keywords)
    companies = (resp or {}).get("companies") or []
    if not companies:
        raise RuntimeError(f"no company matched: {keywords!r}")
    cid = companies[0].get("companyId") or companies[0].get("id")
    if not cid:
        raise RuntimeError(
            f"search_companies returned a company with no id field: {companies[0]!r}")
    return str(cid)


def source_org_followers(client, company_id):
    return (((client.org_followers(company_id) or {}).get("followers")) or [])[:MAX_SOURCE_RESULTS]


def source_group(client, group_id):
    return (((client.group_members(group_id=group_id) or {}).get("members")) or [])[:MAX_SOURCE_RESULTS]


def source_event(client, event_id):
    return (((client.event_attendees(event_id) or {}).get("attendees")) or [])[:MAX_SOURCE_RESULTS]


def dispatch_source(mode, client, **kw):
    if mode == "people":
        return source_people(client, kw["keywords"], filters=kw.get("filters"))
    if mode == "org-followers":
        cid = kw.get("company_id") or resolve_company_id(client, kw["keywords"])
        return source_org_followers(client, cid)
    if mode == "group":
        return source_group(client, kw["group_id"])
    if mode == "event":
        return source_event(client, kw["event_id"])
    raise ValueError(f"unknown mode: {mode}")


# ---------------------------------------------------------------------------
# CLI wiring + end-to-end run
# ---------------------------------------------------------------------------

def parse_args(argv):
    ap = argparse.ArgumentParser(description="LinkedIn-native lead sourcer → Odoo rows")
    ap.add_argument("--mode", required=True,
                    choices=["people", "org-followers", "group", "event"])
    ap.add_argument("--keywords")
    ap.add_argument("--filters")
    ap.add_argument("--company-id")
    ap.add_argument("--group-id")
    ap.add_argument("--event-id")
    ap.add_argument("--exclude-file", help="newline-delimited raw Odoo x_linkedin_url values")
    ap.add_argument("--keyword-score", action="append", default=[],
                    help="ICP keyword (repeatable) for the cheap pre-filter")
    ap.add_argument("--threshold", type=int, default=1)
    ap.add_argument("--cap", type=int, default=120)
    ap.add_argument("--floor", type=int, default=5)
    ap.add_argument("--checkpoint", default="enrich_checkpoint.json")
    ap.add_argument("--schema-manifest",
                    help="JSON list of mailing.contact field names (agent writes it from "
                         "fields_get); when given, verify_schema fails fast before sourcing")
    ap.add_argument("--out", required=True)
    return ap.parse_args(argv)


def _read_exclude_file(path):
    if not path or not os.path.exists(path):
        return []
    with open(path, encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip()]


def run(args, client=None):
    client = client or get_client()
    # Fail fast on a schema drift BEFORE sourcing or spending any enrich budget.
    if getattr(args, "schema_manifest", None) and os.path.exists(args.schema_manifest):
        with open(args.schema_manifest, encoding="utf-8") as f:
            verify_schema(set(json.load(f)))
    raw_urls = _read_exclude_file(args.exclude_file)
    excl_slugs, excl_secondary, dropped = build_exclude_sets(raw_urls)
    if raw_urls and dropped > len(raw_urls) * 0.2:
        print(f"WARNING: {dropped}/{len(raw_urls)} Odoo URLs dropped from exclude set "
              "— dedup coverage degraded.")
    people = dispatch_source(
        args.mode, client, keywords=args.keywords, filters=args.filters,
        company_id=args.company_id, group_id=args.group_id, event_id=args.event_id)
    survivors, dropped_dups = dedup(people, excl_slugs, excl_secondary)
    kept, rejected = score_and_partition(survivors, args.keyword_score, args.threshold)
    enriched, _state = enrich(kept, client, args.checkpoint,
                              cap=args.cap, floor=args.floor)
    leads = []
    for p in enriched:
        leads.append({
            "slug": p.get("slug", ""),
            "profileUrl": p.get("profileUrl", ""),
            "first_name": p.get("firstName", ""),
            "last_name": p.get("lastName", ""),
            "x_headline": p.get("headline", ""),
            "x_job_title": p.get("x_job_title", "") or p.get("currentPosition", ""),
            "company_name": p.get("company_name", ""),
            "x_summary": p.get("x_summary", ""),
            "location": p.get("location", ""),
            "cheap_score": p.get("cheap_score", ""),
            "enriched": p.get("enriched", False),
            # untrusted: an enrich failure carries the API error body — write_workbook
            # neutralizes every cell, so this is the end-to-end injection-safe path.
            "enrich_error": p.get("enrich_error", ""),
            "x_lead_status": "New",
            # not-enriched OR resumed-from-a-prior-run (field-blank here) -> never ready,
            # so a human cannot create a blank/incomplete mailing.contact from this run.
            "odoo_ready": "no" if (not p.get("enriched") or p.get("resumed")) else "",
        })
    reject_rows = [{"slug": r.get("slug", ""),   # use the dedup-time slug, not a recompute
                    "profileUrl": r.get("profileUrl", ""),
                    "first_name": r.get("firstName", ""),
                    "x_headline": r.get("headline", ""),
                    "cheap_score": r.get("cheap_score", 0),
                    "reject_reason": "below ICP threshold"} for r in rejected]
    run_config = {"mode": args.mode, "keywords": args.keywords or "",
                  "threshold": args.threshold, "cap": args.cap,
                  "excluded_existing": len(excl_slugs),
                  "sourced": len(people), "after_dedup": len(survivors),
                  "enriched": sum(1 for p in enriched if p.get("enriched"))}
    write_workbook(args.out, leads, reject_rows, run_config)
    print(f"wrote {len(leads)} leads to {args.out}")
    return args.out


def main(argv=None):
    args = parse_args(argv if argv is not None else sys.argv[1:])
    preflight()
    run(args)


if __name__ == "__main__":
    main()
