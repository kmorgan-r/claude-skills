import contextlib
import io
import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import MagicMock, Mock, patch

import openpyxl

import lead_crawler


class LeadCrawlerTests(unittest.TestCase):
    def test_provider_catalog_has_swappable_search_and_extract_options(self):
        catalog = lead_crawler.provider_catalog()

        self.assertIn("serper", catalog["search"])
        self.assertIn("serpapi", catalog["search"])
        self.assertIn("codex_manual", catalog["search"])
        self.assertIn("codex_builtin", catalog["extract"])
        self.assertIn("jina", catalog["extract"])
        self.assertEqual(catalog["search"]["serper"]["env"], "SERPER_API_KEY")
        self.assertEqual(catalog["extract"]["firecrawl"]["env"], "FIRECRAWL_API_KEY")

    def test_explicit_api_key_takes_precedence_over_environment(self):
        key = lead_crawler.resolve_provider_key(
            "serper",
            lead_crawler.provider_catalog()["search"]["serper"],
            explicit_key="explicit-key",
            env={"SERPER_API_KEY": "env-key"},
        )

        self.assertEqual(key, "explicit-key")

    def test_provider_key_uses_environment_when_no_explicit_key(self):
        key = lead_crawler.resolve_provider_key(
            "serper",
            lead_crawler.provider_catalog()["search"]["serper"],
            explicit_key=None,
            env={"SERPER_API_KEY": "env-key"},
        )

        self.assertEqual(key, "env-key")

    def test_provider_without_key_requirement_does_not_prompt(self):
        prompts = []
        key = lead_crawler.resolve_provider_key(
            "codex_manual",
            lead_crawler.provider_catalog()["search"]["codex_manual"],
            explicit_key=None,
            env={},
            prompt_fn=lambda prompt: prompts.append(prompt) or "prompted-key",
        )

        self.assertIsNone(key)
        self.assertEqual(prompts, [])

    def test_prebuilt_themes_include_taxonomy_and_linkedin_guardrail(self):
        themes = lead_crawler.prebuilt_themes()

        self.assertIn("eu-taxonomy-lca", themes)
        self.assertIn("linkedin-assisted-cross-reference", themes)
        self.assertTrue(themes["linkedin-assisted-cross-reference"]["manual_seed_only"])
        self.assertIn("Product / manufacturing LCA", themes["eu-taxonomy-lca"]["subthemes"])

    def test_query_expansion_uses_theme_terms_and_location(self):
        themes = lead_crawler.prebuilt_themes()
        queries = lead_crawler.expand_queries(
            themes["dpp-rollout-sectors"],
            location="Germany",
            max_queries=4,
        )

        self.assertGreaterEqual(len(queries), 1)
        self.assertLessEqual(len(queries), 4)
        self.assertTrue(any("Germany" in query for query in queries))
        self.assertTrue(any("site:" not in query for query in queries))

    def test_contact_search_queries_target_personas_and_company(self):
        lead = {
            "company_name": "Example Textiles",
            "domain": "example-textiles.com",
            "website": "https://example-textiles.com",
        }
        theme = lead_crawler.prebuilt_themes()["dpp-rollout-sectors"]

        queries = lead_crawler.contact_search_queries(lead, theme, max_queries=5)

        self.assertLessEqual(len(queries), 5)
        self.assertTrue(any('"Example Textiles"' in query for query in queries))
        self.assertTrue(any("site:example-textiles.com" in query for query in queries))
        self.assertTrue(any('"Head of Sustainability"' in query for query in queries))
        self.assertTrue(any('"ESG Manager"' in query for query in queries))

    def test_dedupe_skips_linkedin_and_normalizes_domains(self):
        results = [
            {"title": "A", "link": "https://Example.com/about", "snippet": "textile manufacturer"},
            {"title": "A duplicate", "link": "https://www.example.com/contact", "snippet": "duplicate"},
            {"title": "LinkedIn", "link": "https://www.linkedin.com/company/example", "snippet": "profile"},
        ]
        theme = lead_crawler.prebuilt_themes()["dpp-rollout-sectors"]

        leads = lead_crawler.leads_from_search_results(results, "test query", "dpp-rollout-sectors", theme)

        self.assertEqual(len(leads), 1)
        self.assertEqual(leads[0]["domain"], "example.com")
        self.assertEqual(leads[0]["source_url"], "https://Example.com/about")

    def test_dedupe_collapses_subdomains_to_one_lead_per_registrable_domain(self):
        results = [
            {"title": "Acme UK", "link": "https://www.acme.co.uk/", "snippet": "textile manufacturer"},
            {"title": "Acme Shop", "link": "https://shop.acme.co.uk/products", "snippet": "duplicate company"},
            {"title": "Acme DE", "link": "https://www.acme.de/a", "snippet": "textile manufacturer"},
            {"title": "Acme DE Shop", "link": "https://shop.acme.de/b", "snippet": "duplicate company"},
            {"title": "Other Corp", "link": "https://other.com/", "snippet": "different company"},
        ]
        theme = lead_crawler.prebuilt_themes()["dpp-rollout-sectors"]

        leads = lead_crawler.leads_from_search_results(results, "test query", "dpp-rollout-sectors", theme)

        self.assertEqual(
            [lead_crawler.registrable_domain(lead["source_url"]) for lead in leads],
            ["acme.co.uk", "acme.de", "other.com"],
        )

    def test_private_ip_results_never_become_leads(self):
        results = [
            {"title": "Metadata", "link": "http://169.254.169.254/company", "snippet": "instance metadata"},
            {"title": "Internal", "link": "https://intranet.corp.internal/about", "snippet": "internal host"},
            {"title": "Private", "link": "https://10.0.0.5/admin", "snippet": "private range"},
            {"title": "Loopback", "link": "http://localhost:8080/x", "snippet": "loopback"},
            {"title": "Real Corp", "link": "https://example-textiles.com/about", "snippet": "textile manufacturer"},
        ]
        theme = lead_crawler.prebuilt_themes()["dpp-rollout-sectors"]

        leads = lead_crawler.leads_from_search_results(results, "test query", "dpp-rollout-sectors", theme)

        self.assertEqual([lead["domain"] for lead in leads], ["example-textiles.com"])

    def test_registrable_domain_handles_multi_label_tlds_and_ports(self):
        self.assertEqual(lead_crawler.registrable_domain("https://shop.acme.co.uk/x"), "acme.co.uk")
        self.assertEqual(lead_crawler.registrable_domain("https://www.acme.de/a"), "acme.de")
        self.assertEqual(lead_crawler.registrable_domain("https://deep.sub.acme.com.au"), "acme.com.au")
        self.assertEqual(lead_crawler.registrable_domain("https://acme.com:8443/path"), "acme.com")
        self.assertEqual(lead_crawler.registrable_domain("https://localhost/x"), "localhost")
        self.assertEqual(lead_crawler.registrable_domain(""), "")

    def test_normalized_domain_strips_ports_and_userinfo(self):
        self.assertEqual(lead_crawler.normalized_domain("https://zoominfo.com:443/company"), "zoominfo.com")
        self.assertEqual(lead_crawler.normalized_domain("https://www.Example.COM:8080/about"), "example.com")
        self.assertEqual(lead_crawler.normalized_domain("https://user:pass@acme.com/x"), "acme.com")
        self.assertEqual(lead_crawler.normalized_domain(""), "")

    def test_blocked_domains_not_bypassed_by_explicit_port(self):
        self.assertTrue(lead_crawler.is_blocked_url("https://zoominfo.com:443/company"))
        self.assertTrue(lead_crawler.is_blocked_url("https://de.zoominfo.com:8443/c/x"))
        self.assertFalse(lead_crawler.is_blocked_url("https://example-textiles.com:443/about"))

    def test_blocked_domains_reject_data_vendors_and_directories(self):
        self.assertTrue(lead_crawler.is_blocked_url("https://ensun.io/list"))
        self.assertTrue(lead_crawler.is_blocked_url("https://zoominfo.com/c/somecompany"))
        self.assertTrue(lead_crawler.is_blocked_url("https://www.crunchbase.com/organization/x"))
        self.assertTrue(lead_crawler.is_blocked_url("https://europages.com/en/company"))
        self.assertFalse(lead_crawler.is_blocked_url("https://example-textiles.com/about"))
        self.assertFalse(lead_crawler.is_blocked_url("https://www.sun-garden.de"))

    def test_bad_company_names_rejected(self):
        bad = [
            "The production of textile fabrics in Germany: tradition, innovation an…Storchenwiege GmbH & Co. KG",
            "Top 100 Textile Manufacturing Companies in Germany (2026)",
            "Textile manufacturing Companies in Germany",
            "Setex: Home",
            "Best 50 Furniture Brands - 2025 Guide",
            "",
        ]
        for name in bad:
            self.assertTrue(lead_crawler.looks_like_bad_company_name(name), f"expected bad: {name!r}")

    def test_good_company_names_accepted(self):
        good = [
            "Storchenwiege GmbH & Co. KG",
            "BRANDS Fashion GmbH",
            "LOBERON GmbH",
            "Sun Garden",
            "Apollo Tyres",
            "Vaude Sport",
        ]
        for name in good:
            self.assertFalse(lead_crawler.looks_like_bad_company_name(name), f"expected good: {name!r}")

    def test_leads_from_search_results_skips_listicle_and_blog_titles(self):
        results = [
            {"title": "Top 100 Textile Manufacturing Companies in Germany (2026)", "link": "https://top100.example.com", "snippet": "directory"},
            {"title": "The production of textile fabrics: tradition and innovation", "link": "https://article.example.com", "snippet": "article"},
            {"title": "Storchenwiege GmbH & Co. KG", "link": "https://storchenwiege.de", "snippet": "manufacturer"},
        ]
        theme = lead_crawler.prebuilt_themes()["dpp-rollout-sectors"]

        leads = lead_crawler.leads_from_search_results(results, "test query", "dpp-rollout-sectors", theme)

        domains = {lead["domain"] for lead in leads}
        self.assertNotIn("top100.example.com", domains)
        self.assertNotIn("article.example.com", domains)
        self.assertIn("storchenwiege.de", domains)
        self.assertEqual(len(leads), 1)

    def test_public_emails_filters_placeholder_and_newsletter_emails(self):
        html = """
        Contact us at info@example-textiles.com.
        Demo text: you@company.com
        Newsletter widget: email@newsletter.com
        Placeholder account: user@gmail.com
        """

        emails = lead_crawler.public_emails(html)

        self.assertEqual(emails, ["info@example-textiles.com"])

    def test_find_contact_link_ignores_teamviewer_false_positive(self):
        html = """
        <a href="https://get.teamviewer.com/663n3ee">Remote support</a>
        <a href="/contact">Contact</a>
        """

        contact = lead_crawler.find_contact_link(html, "https://example.com")

        self.assertEqual(contact, "https://example.com/contact")

    def test_extract_contact_people_finds_named_sustainability_contact(self):
        html = """
        <section>
          <h2>Leadership</h2>
          <p>Jane Miller, Head of Sustainability</p>
          <a href="mailto:jane.miller@example-textiles.com">jane.miller@example-textiles.com</a>
        </section>
        """
        theme = lead_crawler.prebuilt_themes()["dpp-rollout-sectors"]

        people = lead_crawler.extract_contact_people(html, "https://example-textiles.com/team", theme)

        self.assertEqual(len(people), 1)
        self.assertEqual(people[0]["contact_name"], "Jane Miller")
        self.assertEqual(people[0]["contact_title"], "Head of Sustainability")
        self.assertEqual(people[0]["contact_email"], "jane.miller@example-textiles.com")
        self.assertEqual(people[0]["contact_source_url"], "https://example-textiles.com/team")
        self.assertGreaterEqual(people[0]["contact_confidence"], 70)

    def test_extract_contact_people_ignores_generic_topic_labels(self):
        html = "<p>Carbon Accounting, ESG Reporting</p>"

        people = lead_crawler.extract_contact_people(html, "https://example.com/blog", {})

        self.assertEqual(people, [])

    def test_extract_contact_people_ignores_marketing_sentence_fragments(self):
        html = "<p>APAC. Real-time tracking of ESG performance is now available.</p>"

        people = lead_crawler.extract_contact_people(html, "https://example.com/blog", {})

        self.assertEqual(people, [])

    def test_extract_contact_people_ignores_masked_directory_snippets(self):
        html = "Audit Manager, ESG Coordinator. Email ****** @****.com. Phone (***) ****-****."

        people = lead_crawler.extract_contact_people(html, "https://directory.example/person", {})

        self.assertEqual(people, [])

    def test_best_email_for_name_matches_full_name_local_part(self):
        emails = ["info@acme.com", "jane.miller@acme.com"]

        self.assertEqual(
            lead_crawler.best_email_for_name("Jane Miller", emails),
            "jane.miller@acme.com",
        )

    def test_best_email_for_name_matches_any_name_part(self):
        emails = ["info@acme.com", "miller@acme.com"]

        self.assertEqual(
            lead_crawler.best_email_for_name("Jane Miller", emails),
            "miller@acme.com",
        )

    def test_best_email_for_name_returns_empty_when_no_match(self):
        # emails[0] here belongs to someone else; attributing it to an
        # unmatched person would send outreach to the wrong inbox.
        emails = ["bob.schmidt@acme.com", "office@acme.com"]

        self.assertEqual(lead_crawler.best_email_for_name("Jane Miller", emails), "")

    def test_best_email_for_name_handles_degenerate_names(self):
        emails = ["info@acme.com"]

        self.assertEqual(lead_crawler.best_email_for_name("", emails), "")
        self.assertEqual(lead_crawler.best_email_for_name("J X", emails), "")

    def test_extract_contact_people_does_not_misattribute_emails_on_team_page(self):
        html = """
        <section>
          <p>Jane Miller, Head of Sustainability</p>
          <p>Bob Schmidt, Sustainability Manager</p>
          <a href="mailto:bob.schmidt@acme.com">bob.schmidt@acme.com</a>
        </section>
        """
        theme = lead_crawler.prebuilt_themes()["dpp-rollout-sectors"]

        people = lead_crawler.extract_contact_people(html, "https://acme.com/team", theme)

        by_name = {person["contact_name"]: person["contact_email"] for person in people}
        self.assertEqual(by_name["Bob Schmidt"], "bob.schmidt@acme.com")
        self.assertEqual(by_name["Jane Miller"], "")

    def test_list_providers_warns_about_codex_builtin_in_cloud(self):
        stdout = io.StringIO()
        with contextlib.redirect_stdout(stdout):
            lead_crawler.list_providers()

        output = stdout.getvalue()
        self.assertIn("codex_builtin", output)
        self.assertIn("WARNING:", output)
        self.assertIn("cloud", output)
        self.assertIn("DNS rebinding", output)

    def test_enrich_public_pages_follows_team_page_for_named_contacts(self):
        theme = lead_crawler.prebuilt_themes()["dpp-rollout-sectors"]
        leads = lead_crawler.leads_from_search_results(
            [
                {
                    "title": "Example Textiles",
                    "link": "https://example-textiles.com",
                    "snippet": "Textile manufacturer",
                }
            ],
            "test query",
            "dpp-rollout-sectors",
            theme,
        )
        pages = {
            "https://example-textiles.com": {
                "url": "https://example-textiles.com",
                "html": '<a href="/team">Team</a><a href="/contact">Contact</a>',
                "text": "Home",
            },
            "https://example-textiles.com/team": {
                "url": "https://example-textiles.com/team",
                "html": """
                <p>Jane Miller, Head of Sustainability</p>
                <a href="mailto:jane.miller@example-textiles.com">Email</a>
                """,
                "text": "Jane Miller, Head of Sustainability",
            },
            "https://example-textiles.com/contact": {
                "url": "https://example-textiles.com/contact",
                "html": '<a href="mailto:info@example-textiles.com">Email us</a>',
                "text": "Contact",
            },
        }

        with patch.object(lead_crawler, "extract_page", side_effect=lambda url, *_: pages[url]):
            lead_crawler.enrich_public_pages(leads, theme, "codex_builtin")

        self.assertEqual(leads[0]["contact_name"], "Jane Miller")
        self.assertEqual(leads[0]["contact_title"], "Head of Sustainability")
        self.assertEqual(leads[0]["contact_email"], "jane.miller@example-textiles.com")
        self.assertEqual(leads[0]["contact_source_url"], "https://example-textiles.com/team")
        self.assertEqual(leads[0]["contact_data_type"], "person")

    def test_enrich_public_pages_outer_failure_preserves_existing_notes(self):
        theme = lead_crawler.prebuilt_themes()["dpp-rollout-sectors"]
        leads = [
            {
                "company_name": "Example Textiles",
                "website": "https://example-textiles.com",
                "notes": "Earlier pipeline note",
            }
        ]

        with patch.object(lead_crawler, "extract_page", side_effect=RuntimeError("primary fetch boom")):
            lead_crawler.enrich_public_pages(leads, theme, "codex_builtin")

        self.assertIn("Earlier pipeline note", leads[0]["notes"])
        self.assertIn("Page crawl failed: primary fetch boom", leads[0]["notes"])

    def test_fixture_export_creates_expected_workbook_sheets_and_columns(self):
        fixture = {
            "organic_results": [
                {
                    "title": "Example Textiles",
                    "link": "https://example-textiles.com/sustainability",
                    "snippet": "Apparel manufacturer publishing product carbon footprint details.",
                }
            ]
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            fixture_path = Path(temp_dir) / "fixture.json"
            output_path = Path(temp_dir) / "leads.xlsx"
            fixture_path.write_text(json.dumps(fixture), encoding="utf-8")

            args = lead_crawler.parse_args(
                [
                    "--theme",
                    "dpp-rollout-sectors",
                    "--fixture",
                    str(fixture_path),
                    "--output",
                    str(output_path),
                    "--no-crawl-pages",
                    "--search-provider",
                    "serper",
                    "--extract-provider",
                    "codex_builtin",
                ]
            )
            lead_crawler.run(args)

            workbook = openpyxl.load_workbook(output_path, data_only=True)
            self.assertEqual(
                workbook.sheetnames,
                ["Leads", "Sources", "Rejected", "Run Config"],
            )
            headers = [cell.value for cell in workbook["Leads"][1]]
            self.assertIn("company_name", headers)
            self.assertIn("linkedin_reference_url", headers)
            self.assertIn("outreach_allowed_review", headers)
            self.assertIn("target_persona", headers)
            self.assertIn("contact_name", headers)
            self.assertIn("contact_title", headers)
            self.assertIn("contact_source_url", headers)
            self.assertIn("contact_confidence", headers)
            self.assertIn("person_source_type", headers)
            self.assertIn("public_profile_url", headers)
            self.assertIn("email_discovery_method", headers)
            self.assertIn("email_verification_status", headers)
            self.assertIn("email_confidence", headers)
            self.assertIn("do_not_contact_reason", headers)
            self.assertEqual(workbook["Leads"].max_row, 2)
            config = {row[0].value: row[1].value for row in workbook["Run Config"].iter_rows(min_row=2)}
            self.assertEqual(config["search_provider"], "fixture")
            self.assertEqual(config["extract_provider"], "codex_builtin")


    def test_fixture_run_with_contact_search_skips_enrichment_without_error(self):
        fixture = {
            "organic_results": [
                {
                    "title": "Example Textiles",
                    "link": "https://example-textiles.com/sustainability",
                    "snippet": "Apparel manufacturer publishing product carbon footprint details.",
                }
            ]
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            fixture_path = Path(temp_dir) / "fixture.json"
            output_path = Path(temp_dir) / "leads.xlsx"
            fixture_path.write_text(json.dumps(fixture), encoding="utf-8")

            args = lead_crawler.parse_args(
                [
                    "--theme",
                    "dpp-rollout-sectors",
                    "--fixture",
                    str(fixture_path),
                    "--output",
                    str(output_path),
                    "--no-crawl-pages",
                    "--contact-search",
                    "--search-provider",
                    "serper",
                    "--extract-provider",
                    "codex_builtin",
                ]
            )
            # Must not raise (search_api_key is never resolved for fixture runs)
            # and must not attempt contact-search enrichment.
            with patch.object(
                lead_crawler, "enrich_contacts_via_search", side_effect=AssertionError("contact search must not run for fixture provider")
            ):
                lead_crawler.run(args)

            workbook = openpyxl.load_workbook(output_path, data_only=True)
            source_providers = [row[2].value for row in workbook["Sources"].iter_rows(min_row=2)]
            self.assertEqual(source_providers, [str(fixture_path)])

    def test_write_table_neutralizes_formula_starters(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        lead_crawler.write_table(
            sheet,
            ["col_a", "col_b"],
            [
                ["=HYPERLINK('http://attacker.com','click')", "normal"],
                ["+1+1", "-2"],
                ["@SUM(A1:A2)", 42],
            ],
        )
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        self.assertEqual(rows[0][0], "'=HYPERLINK('http://attacker.com','click')")
        self.assertEqual(rows[0][1], "normal")
        self.assertEqual(rows[1][0], "'+1+1")
        self.assertEqual(rows[1][1], "'-2")
        self.assertEqual(rows[2][0], "'@SUM(A1:A2)")
        self.assertEqual(rows[2][1], 42)

    def test_fetch_text_rejects_non_http_scheme(self):
        with patch.object(lead_crawler, "requests") as mock_requests:
            result = lead_crawler.fetch_text("ftp://example.com/file.txt")
            self.assertEqual(result, "")
            mock_requests.get.assert_not_called()

    def test_fetch_text_rejects_private_ips(self):
        with patch.object(lead_crawler, "requests") as mock_requests:
            for bad_url in [
                "http://192.168.1.1/admin",
                "http://10.0.0.1/config",
                "http://127.0.0.1/secret",
                "http://169.254.169.254/latest/meta-data/",
                "http://172.16.0.1/api",
                "http://[::1]/admin",
                # 0.0.0.0 routes to loopback on Linux; on Python <= 3.10 only
                # is_unspecified flags it (is_private gained 0.0.0.0/8 in 3.11).
                "http://0.0.0.0/admin",
                "http://0.0.0.0:8080/internal",
                "http://[::]/admin",
            ]:
                result = lead_crawler.fetch_text(bad_url)
                self.assertEqual(result, "", f"expected empty for {bad_url}")
            mock_requests.get.assert_not_called()

    def test_candidate_contact_links_rejects_private_ip_urls(self):
        html = '<a href="http://192.168.1.1/contact">Contact</a><a href="http://192.168.1.1/about">About</a>'
        links = lead_crawler.candidate_contact_links(html, "http://192.168.1.1")
        self.assertEqual(links, [])


    def test_is_private_ip_url_blocks_hostname_internal_targets(self):
        blocked = [
            "http://localhost/admin",
            "http://localhost.localdomain:8080/",
            "https://metadata.google.internal/computeMetadata/v1/",
            "http://metadata:80/",
            "http://internal-service.local/path",
            "https://anything.internal/",
        ]
        for url in blocked:
            self.assertTrue(lead_crawler._is_private_ip_url(url), f"expected blocked: {url}")

        public = [
            "https://example.com/about",
            "http://www.google.com/search",
            "https://api.serper.dev/search",
        ]
        for url in public:
            self.assertFalse(lead_crawler._is_private_ip_url(url), f"expected public: {url}")

    def test_jina_extract_rejects_bad_scheme_and_private_ip(self):
        with patch.object(lead_crawler, "requests") as mock_requests:
            result = lead_crawler.jina_extract("ftp://example.com/file", None)
            self.assertEqual(result, {"url": "ftp://example.com/file", "text": "", "emails": []})
            mock_requests.get.assert_not_called()

            result = lead_crawler.jina_extract("http://localhost/secret", "key")
            self.assertEqual(result, {"url": "http://localhost/secret", "text": "", "emails": []})
            mock_requests.get.assert_not_called()

            mock_requests.get.return_value.text = "Extracted text"
            lead_crawler.jina_extract("https://example.com/page?id=1", "key")
            call_url = mock_requests.get.call_args[0][0]
            self.assertTrue(call_url.startswith("https://r.jina.ai/"))
            self.assertIn("https%3A%2F%2Fexample.com%2Fpage%3Fid%3D1", call_url)

    def test_read_manual_seeds_normalizes_record_shapes(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            seed_path = Path(temp_dir) / "seeds.json"
            seed_path.write_text(
                json.dumps([
                    {"company": "Good Corp", "url": "https://good.example.com"},
                    {"company": "Ref Corp", "linkedin": "https://www.linkedin.com/company/ref-corp"},
                    "acme.de",
                    "https://www.linkedin.com/company/acme",
                    "Name Only Corp",
                    "   ",
                ]),
                encoding="utf-8",
            )
            records = lead_crawler.read_manual_seeds(str(seed_path))

        by_company = {r["company"]: r for r in records}
        # Website seed kept as a crawlable site.
        self.assertEqual(by_company["Good Corp"]["website"], "https://good.example.com")
        # LinkedIn URL never becomes a crawl target; stored as a reference.
        self.assertEqual(by_company["Ref Corp"]["website"], "")
        self.assertIn("linkedin.com/company/ref-corp", by_company["Ref Corp"]["linkedin"])
        # Bare domain gains a scheme.
        self.assertEqual(by_company[""]["website"], "https://acme.de")
        # Bare LinkedIn URL -> company name derived from the slug, no website.
        self.assertEqual(by_company["Acme"]["website"], "")
        self.assertIn("linkedin.com/company/acme", by_company["Acme"]["linkedin"])
        # Name-only seed is kept (company list input), no website.
        self.assertEqual(by_company["Name Only Corp"]["website"], "")
        # Blank entry dropped.
        self.assertNotIn("   ", by_company)

    def test_read_manual_seeds_parses_csv(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            seed_path = Path(temp_dir) / "seeds.csv"
            seed_path.write_text(
                "company,url,linkedin\n"
                "Acme GmbH,acme.de,https://www.linkedin.com/company/acme\n"
                "Bravo Ltd,https://bravo.example.com,\n",
                encoding="utf-8",
            )
            records = lead_crawler.read_manual_seeds(str(seed_path))

        by_company = {r["company"]: r for r in records}
        self.assertEqual(by_company["Acme GmbH"]["website"], "https://acme.de")
        self.assertIn("linkedin.com/company/acme", by_company["Acme GmbH"]["linkedin"])
        self.assertEqual(by_company["Bravo Ltd"]["website"], "https://bravo.example.com")

    def test_leads_from_manual_seeds_creates_linkedin_reference_rows(self):
        theme = lead_crawler.prebuilt_themes()["linkedin-assisted-cross-reference"]
        records = [
            {"company": "Acme", "website": "", "linkedin": "https://www.linkedin.com/company/acme"},
            {"company": "Bravo", "website": "https://bravo.com", "linkedin": ""},
        ]
        leads = lead_crawler.leads_from_manual_seeds(records, "linkedin-assisted-cross-reference", theme, None)
        # Only the website-less seed becomes a reference row here.
        self.assertEqual(len(leads), 1)
        self.assertEqual(leads[0]["company_name"], "Acme")
        self.assertIn("linkedin.com/company/acme", leads[0]["linkedin_reference_url"])
        self.assertEqual(leads[0]["person_source_type"], "linkedin_reference")
        self.assertEqual(leads[0]["contact_data_type"], "company")

    def test_seeds_as_search_items_only_crawlable(self):
        records = [
            {"company": "Acme", "website": "https://acme.com", "linkedin": "https://linkedin.com/company/acme"},
            {"company": "Ref", "website": "", "linkedin": "https://linkedin.com/company/ref"},
        ]
        items = lead_crawler.seeds_as_search_items(records)
        self.assertEqual(len(items), 1)
        self.assertEqual(items[0]["link"], "https://acme.com")
        self.assertEqual(items[0]["linkedin_reference_url"], "https://linkedin.com/company/acme")

    def test_enrich_contacts_via_search_respects_budget(self):
        leads = [
            {"company_name": "A", "domain": "a.com"},
            {"company_name": "B", "domain": "b.com"},
        ]
        theme = lead_crawler.prebuilt_themes()["dpp-rollout-sectors"]

        with patch.object(lead_crawler, "contact_search_queries", return_value=["q1", "q2"]), \
             patch.object(lead_crawler, "search_provider", return_value=[{"link": "https://a.com"}]):
            sources = lead_crawler.enrich_contacts_via_search(
                leads, "serper", "key", theme, per_lead_queries=2, budget=1
            )
        self.assertEqual(len(sources), 1)

    def test_enrich_contacts_via_search_survives_outer_exception(self):
        leads = [
            {"company_name": "A", "domain": "a.com"},
            {"company_name": "B", "domain": "b.com"},
        ]
        theme = lead_crawler.prebuilt_themes()["dpp-rollout-sectors"]

        def side_effect(lead, theme, max_queries):
            if lead["company_name"] == "A":
                raise RuntimeError("boom")
            return ["q1"]

        with patch.object(lead_crawler, "contact_search_queries", side_effect=side_effect), \
             patch.object(lead_crawler, "search_provider", return_value=[{"link": "https://b.com"}]):
            sources = lead_crawler.enrich_contacts_via_search(
                leads, "serper", "key", theme, per_lead_queries=1, budget=0
            )
        self.assertEqual(len(sources), 1)
        self.assertEqual(leads[0].get("notes"), "Contact search interrupted: boom")

    @staticmethod
    def _getaddrinfo_result(ip: str) -> list[tuple]:
        return [(2, 1, 6, "", (ip, 0))]

    @staticmethod
    def _fetch_response(body: str = "", *, is_redirect: bool = False, location: str = "", encoding: str = "utf-8"):
        # fetch_text uses `with requests.get(...) as response` and reads
        # response.raw, so the mock must be a context manager exposing raw.read.
        resp = MagicMock()
        resp.is_redirect = is_redirect
        resp.raise_for_status.return_value = None
        resp.headers = {"Location": location} if location else {}
        resp.encoding = encoding
        resp.apparent_encoding = encoding
        resp.raw.read.return_value = body.encode(encoding)
        resp.__enter__.return_value = resp
        resp.__exit__.return_value = False
        return resp

    def test_fetch_text_follows_safe_redirect(self):
        with patch.object(lead_crawler, "requests") as mock_requests, \
             patch.object(lead_crawler.socket, "getaddrinfo", return_value=self._getaddrinfo_result("93.184.216.34")):
            redirect_resp = self._fetch_response(is_redirect=True, location="https://example.com/")
            ok_resp = self._fetch_response("redirected content")

            mock_requests.get.side_effect = [redirect_resp, ok_resp]
            result = lead_crawler.fetch_text("http://example.com")
            self.assertEqual(result, "redirected content")
            self.assertEqual(mock_requests.get.call_count, 2)

    def test_fetch_text_blocks_redirect_to_private_ip(self):
        with patch.object(lead_crawler, "requests") as mock_requests, \
             patch.object(lead_crawler.socket, "getaddrinfo", return_value=self._getaddrinfo_result("93.184.216.34")):
            redirect_resp = self._fetch_response(is_redirect=True, location="http://192.168.1.1/secret")

            mock_requests.get.return_value = redirect_resp
            result = lead_crawler.fetch_text("http://example.com")
            self.assertEqual(result, "")
            mock_requests.get.assert_called_once()

    def test_fetch_text_blocks_redirect_to_blocked_domain(self):
        # A lead site that 301s to a no-scrape domain (LinkedIn) must not be
        # followed, even though linkedin.com is a public host.
        with patch.object(lead_crawler, "requests") as mock_requests, \
             patch.object(lead_crawler.socket, "getaddrinfo", return_value=self._getaddrinfo_result("93.184.216.34")):
            redirect_resp = self._fetch_response(is_redirect=True, location="https://www.linkedin.com/company/acme")
            mock_requests.get.return_value = redirect_resp
            result = lead_crawler.fetch_text("https://acme.com")
            self.assertEqual(result, "")
            mock_requests.get.assert_called_once()

    def test_fetch_text_decodes_charsetless_utf8_pages(self):
        # text/html without a charset: requests defaults encoding to
        # ISO-8859-1, which would mojibake UTF-8 names. fetch_text must fall
        # back to the sniffed encoding instead.
        with patch.object(lead_crawler, "requests") as mock_requests, \
             patch.object(lead_crawler.socket, "getaddrinfo", return_value=self._getaddrinfo_result("93.184.216.34")):
            resp = MagicMock()
            resp.is_redirect = False
            resp.raise_for_status.return_value = None
            resp.headers = {}
            resp.encoding = "ISO-8859-1"
            resp.apparent_encoding = "utf-8"
            resp.raw.read.return_value = "Jürgen Müller".encode("utf-8")
            resp.__enter__.return_value = resp
            resp.__exit__.return_value = False
            mock_requests.get.return_value = resp
            result = lead_crawler.fetch_text("https://example.com/team")
            self.assertEqual(result, "Jürgen Müller")

    def test_fetch_text_refuses_hostname_resolving_to_private_ip(self):
        # DNS rebinding: hostname looks public but resolves to an internal or
        # cloud-metadata address. Must be refused before any connection.
        for rebound_ip in ["169.254.169.254", "10.0.0.5", "127.0.0.1", "::1", "0.0.0.0", "::"]:
            with patch.object(lead_crawler, "requests") as mock_requests, \
                 patch.object(lead_crawler.socket, "getaddrinfo", return_value=self._getaddrinfo_result(rebound_ip)):
                result = lead_crawler.fetch_text("https://attacker-seo-domain.com/")
                self.assertEqual(result, "", f"expected refusal for hostname resolving to {rebound_ip}")
                mock_requests.get.assert_not_called()

    def test_fetch_text_blocks_redirect_to_hostname_resolving_private(self):
        def fake_getaddrinfo(hostname, *args, **kwargs):
            if hostname == "rebind.example.net":
                return self._getaddrinfo_result("169.254.169.254")
            return self._getaddrinfo_result("93.184.216.34")

        with patch.object(lead_crawler, "requests") as mock_requests, \
             patch.object(lead_crawler.socket, "getaddrinfo", side_effect=fake_getaddrinfo):
            redirect_resp = self._fetch_response(is_redirect=True, location="https://rebind.example.net/meta")

            mock_requests.get.return_value = redirect_resp
            result = lead_crawler.fetch_text("http://example.com")
            self.assertEqual(result, "")
            mock_requests.get.assert_called_once()

    def test_fetch_text_proceeds_when_hostname_unresolvable(self):
        # Fail-open on resolution errors: the request itself would fail the
        # same way for a truly unresolvable host, so nothing can leak.
        with patch.object(lead_crawler, "requests") as mock_requests, \
             patch.object(lead_crawler.socket, "getaddrinfo", side_effect=lead_crawler.socket.gaierror("no such host")):
            ok_resp = self._fetch_response("content")

            mock_requests.get.return_value = ok_resp
            result = lead_crawler.fetch_text("https://example.com/")
            self.assertEqual(result, "content")
            mock_requests.get.assert_called_once()

    def test_firecrawl_extract_rejects_bad_scheme_and_private_ip(self):
        with patch.object(lead_crawler, "requests") as mock_requests:
            result = lead_crawler.firecrawl_extract("ftp://example.com/file", "key")
            self.assertEqual(result, {"url": "ftp://example.com/file", "text": "", "html": "", "emails": []})
            mock_requests.post.assert_not_called()

            result = lead_crawler.firecrawl_extract("http://localhost/secret", "key")
            self.assertEqual(result, {"url": "http://localhost/secret", "text": "", "html": "", "emails": []})
            mock_requests.post.assert_not_called()

    def test_tavily_extract_rejects_bad_scheme_and_private_ip(self):
        with patch.object(lead_crawler, "requests") as mock_requests:
            result = lead_crawler.tavily_extract("ftp://example.com/file", "key")
            self.assertEqual(result, {"url": "ftp://example.com/file", "text": "", "emails": []})
            mock_requests.post.assert_not_called()

            result = lead_crawler.tavily_extract("http://localhost/secret", "key")
            self.assertEqual(result, {"url": "http://localhost/secret", "text": "", "emails": []})
            mock_requests.post.assert_not_called()

    def test_exa_extract_rejects_bad_scheme_and_private_ip(self):
        with patch.object(lead_crawler, "requests") as mock_requests:
            result = lead_crawler.exa_extract("ftp://example.com/file", "key")
            self.assertEqual(result, {"url": "ftp://example.com/file", "text": "", "emails": []})
            mock_requests.post.assert_not_called()

            result = lead_crawler.exa_extract("http://localhost/secret", "key")
            self.assertEqual(result, {"url": "http://localhost/secret", "text": "", "emails": []})
            mock_requests.post.assert_not_called()

    def test_contact_search_queries_survives_empty_titles(self):
        lead = {
            "company_name": "Example Textiles",
            "website": "https://example-textiles.com",
        }
        theme = {"contact_search_titles": []}

        queries = lead_crawler.contact_search_queries(lead, theme, max_queries=5)

        self.assertEqual(queries, [])

    def test_enrich_contacts_via_search_redacts_api_key_in_notes(self):
        leads = [{"company_name": "A", "domain": "a.com"}]
        theme = lead_crawler.prebuilt_themes()["dpp-rollout-sectors"]
        secret_key = "sk_live_secret_12345"

        def side_effect(query, provider, key, max_results):
            raise RuntimeError(f"403 for url: https://serpapi.com/search?api_key={key}")

        with patch.object(lead_crawler, "contact_search_queries", return_value=["q1"]), \
             patch.object(lead_crawler, "search_provider", side_effect=side_effect):
            lead_crawler.enrich_contacts_via_search(
                leads, "serpapi", secret_key, theme, per_lead_queries=1, budget=0
            )
        self.assertNotIn(secret_key, leads[0].get("notes", ""))
        self.assertIn("***", leads[0].get("notes", ""))

    def test_read_fixture_supports_multiple_provider_formats(self):
        item = {"title": "Example", "link": "https://example.com", "snippet": "s"}
        payloads = [
            {"organic_results": [item]},  # SerpApi / SearchApi
            {"organic": [item]},  # Serper
            {"results": [item]},  # Tavily
            [item],  # bare list
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            for index, payload in enumerate(payloads):
                fixture_path = Path(temp_dir) / f"fixture_{index}.json"
                fixture_path.write_text(json.dumps(payload), encoding="utf-8")
                self.assertEqual(
                    lead_crawler.read_fixture(str(fixture_path)),
                    [item],
                    msg=f"payload format {index} should yield results",
                )

    def test_run_survives_failed_query_and_redacts_key_in_sources(self):
        secret_key = "sk_live_secret_12345"
        calls = {"count": 0}

        def side_effect(query, provider_id, api_key, max_results):
            calls["count"] += 1
            if calls["count"] == 1:
                raise RuntimeError(
                    f"429 Too Many Requests for url: https://serpapi.com/search.json?api_key={api_key}"
                )
            return [
                {
                    "title": "Example Textiles",
                    "link": "https://example-textiles.com/sustainability",
                    "snippet": "Apparel manufacturer publishing product carbon footprint details.",
                }
            ]

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "leads.xlsx"
            args = lead_crawler.parse_args(
                [
                    "--theme",
                    "dpp-rollout-sectors",
                    "--search-provider",
                    "serpapi",
                    "--search-api-key",
                    secret_key,
                    "--extract-provider",
                    "codex_builtin",
                    "--no-crawl-pages",
                    "--max-queries",
                    "2",
                    "--output",
                    str(output_path),
                ]
            )
            with patch.object(lead_crawler, "search_provider", side_effect=side_effect):
                lead_crawler.run(args)

            workbook = openpyxl.load_workbook(output_path, data_only=True)
            self.assertEqual(workbook["Leads"].max_row, 2)
            source_rows = list(workbook["Sources"].iter_rows(min_row=2, values_only=True))
            error_rows = [row for row in source_rows if str(row[2]).startswith("error:")]
            self.assertEqual(len(error_rows), 1)
            self.assertNotIn(secret_key, str(error_rows[0][2]))
            self.assertIn("***", str(error_rows[0][2]))
            ok_rows = [row for row in source_rows if row[2] == "serpapi"]
            self.assertEqual(len(ok_rows), 1)

    def test_raise_for_status_redacted_strips_key_from_http_error(self):
        secret_key = "sk_live_secret_12345"
        response = Mock()
        response.raise_for_status.side_effect = lead_crawler.requests.HTTPError(
            f"400 Client Error: Bad Request for url: https://serpapi.com/search.json?q=x&api_key={secret_key}"
        )

        with self.assertRaises(lead_crawler.requests.HTTPError) as ctx:
            lead_crawler._raise_for_status_redacted(response, secret_key)

        self.assertNotIn(secret_key, str(ctx.exception))
        self.assertIn("***", str(ctx.exception))

    def test_extract_contact_people_handles_diacritics_and_dashes(self):
        theme = lead_crawler.prebuilt_themes()["dpp-rollout-sectors"]
        html = """
        <p>Jürgen Müller, Head of Sustainability</p>
        <p>Anna Larsen – Sustainability Manager</p>
        """
        people = lead_crawler.extract_contact_people(html, "https://acme.de/team", theme)
        names = {p["contact_name"] for p in people}
        self.assertIn("Jürgen Müller", names)
        self.assertIn("Anna Larsen", names)

    def test_clean_company_name_splits_en_dash(self):
        self.assertEqual(
            lead_crawler.clean_company_name("Möbelwerk Huber GmbH – Nachhaltige Möbel", "huber.de"),
            "Möbelwerk Huber GmbH",
        )

    def test_leads_from_search_results_reads_content_key(self):
        theme = lead_crawler.prebuilt_themes()["dpp-rollout-sectors"]
        # Tavily/Exa-shaped item: text under 'content', url under 'url'.
        results = [{"title": "Acme", "url": "https://acme.com", "content": "ISO 14067 product carbon footprint"}]
        leads = lead_crawler.leads_from_search_results(results, "q", "dpp-rollout-sectors", theme)
        self.assertEqual(leads[0]["evidence_snippet"], "ISO 14067 product carbon footprint")
        self.assertGreater(leads[0]["lead_score"], 30)

    def test_per_lead_query_attribution(self):
        theme = lead_crawler.prebuilt_themes()["dpp-rollout-sectors"]
        results = [
            {"title": "Acme", "link": "https://acme.com", "snippet": "m", "_query": "query-A"},
            {"title": "Bravo", "link": "https://bravo.com", "snippet": "m", "_query": "query-B"},
        ]
        leads = lead_crawler.leads_from_search_results(results, "joined | queries", "dpp-rollout-sectors", theme)
        basis = {lead["company_name"]: lead["business_relevance_basis"] for lead in leads}
        self.assertEqual(basis["Acme"], "query-A")
        self.assertEqual(basis["Bravo"], "query-B")

    def test_apply_page_enrichment_sets_provenance_fields(self):
        theme = lead_crawler.prebuilt_themes()["dpp-rollout-sectors"]
        lead = lead_crawler.new_lead(company_name="Acme", domain="acme.de", website="https://acme.de")
        page = {
            "url": "https://acme.de/team",
            "html": '<p>Jane Miller, Head of Sustainability</p><a href="mailto:jane.miller@acme.de">x</a>',
            "text": "Jane Miller, Head of Sustainability",
        }
        lead_crawler.apply_page_enrichment(lead, page, theme)
        self.assertEqual(lead["contact_data_type"], "person")
        self.assertEqual(lead["person_source_type"], "company_page")
        self.assertEqual(lead["email_discovery_method"], "public_page")
        self.assertEqual(lead["email_verification_status"], "unverified")

    def test_candidate_contact_links_follows_markdown_links(self):
        markdown = "Welcome. [Our Team](/team) and [Contact us](/contact)."
        links = lead_crawler.candidate_contact_links(markdown, "https://acme.de")
        self.assertIn("https://acme.de/team", links)
        self.assertIn("https://acme.de/contact", links)

    def test_merge_linkedin_references_matches_cleaned_name(self):
        theme = lead_crawler.prebuilt_themes()["dpp-rollout-sectors"]
        records = [
            {"company": "Acme GmbH - Official Site", "website": "https://acme.de",
             "linkedin": "https://www.linkedin.com/company/acme"}
        ]
        leads = lead_crawler.leads_from_search_results(
            [{"title": "Acme GmbH - Official Site", "link": "https://acme.de", "snippet": "m"}],
            "q", "dpp-rollout-sectors", theme,
        )
        self.assertEqual(leads[0]["company_name"], "Acme GmbH")
        lead_crawler.merge_linkedin_references(leads, records)
        self.assertIn("linkedin.com/company/acme", leads[0]["linkedin_reference_url"])

    def test_region_for_location_classifies_countries(self):
        self.assertEqual(lead_crawler.region_for_location("United States")["region"], "US")
        self.assertEqual(lead_crawler.region_for_location("us")["region"], "US")
        germany = lead_crawler.region_for_location("Munich, Germany")
        self.assertEqual(germany["region"], "EU")
        self.assertTrue(germany["strict"])  # UWG strict
        self.assertEqual(lead_crawler.region_for_location("DE")["country"], "Germany")
        self.assertEqual(lead_crawler.region_for_location("France")["region"], "EU")
        self.assertFalse(lead_crawler.region_for_location("France")["strict"])
        self.assertEqual(lead_crawler.region_for_location("")["region"], "EU")
        self.assertEqual(lead_crawler.region_for_location("Brazil")["region"], "unknown")

    def test_compliance_fields_posture_by_region(self):
        us = lead_crawler.compliance_fields({"region": "US"})
        self.assertEqual(us["consent_status"], "n/a (opt-out)")
        self.assertIn("CAN-SPAM", us["outreach_allowed_review"])

        de = lead_crawler.compliance_fields({"region": "EU", "strict": True})
        self.assertEqual(de["consent_status"], "unknown")
        self.assertEqual(de["outreach_allowed_review"], "needs review")
        self.assertIn("UWG", de["legitimate_interest_basis"])

        unknown = lead_crawler.compliance_fields({"region": "unknown"})
        self.assertEqual(unknown["outreach_allowed_review"], "needs review")
        self.assertNotIn("UWG", unknown["legitimate_interest_basis"])

    def test_leads_carry_region_aware_compliance_posture(self):
        theme = lead_crawler.prebuilt_themes()["dpp-rollout-sectors"]
        results = [{"title": "Acme Inc", "link": "https://acme.com", "snippet": "Manufacturer"}]
        leads = lead_crawler.leads_from_search_results(
            results, "q", "dpp-rollout-sectors", theme, lead_crawler.region_for_location("United States")
        )
        self.assertEqual(leads[0]["region"], "US")
        self.assertEqual(leads[0]["consent_status"], "n/a (opt-out)")
        self.assertEqual(leads[0]["country"], "United States")

    def test_contact_provenance_ok_requires_company_domain(self):
        self.assertTrue(
            lead_crawler.contact_provenance_ok("https://acme.com/team", "https://acme.com")
        )
        self.assertTrue(
            lead_crawler.contact_provenance_ok("https://www.acme.com/about/jane", "https://acme.com")
        )
        # Data-vendor / third-party page: not acceptable provenance.
        self.assertFalse(
            lead_crawler.contact_provenance_ok("https://theorg.com/org/acme/jane", "https://acme.com")
        )
        self.assertFalse(lead_crawler.contact_provenance_ok("", "https://acme.com"))

    def test_apply_contact_search_results_rejects_offdomain_evidence(self):
        theme = lead_crawler.prebuilt_themes()["dpp-rollout-sectors"]
        lead = {"company_name": "Acme", "domain": "acme.com", "website": "https://acme.com"}
        results = [
            {
                "title": "Jane Miller, Head of Sustainability",
                "link": "https://datanyze.com/companies/acme/jane-miller",
                "snippet": "Jane Miller, Head of Sustainability at Acme",
            }
        ]
        lead_crawler.apply_contact_search_results(lead, results, "q", theme)
        # Off-domain vendor snippet must not promote the lead to a named person.
        self.assertNotEqual(lead.get("contact_data_type"), "person")
        self.assertEqual(lead.get("contact_name", ""), "")

    def test_write_table_strips_xml_illegal_characters(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            out = Path(temp_dir) / "out.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            # \x0c (form feed) is XML-illegal; openpyxl raises on save without scrubbing.
            lead_crawler.write_table(sheet, ["col"], [["clean\x0ctext"]])
            workbook.save(out)
            reopened = openpyxl.load_workbook(out)
            self.assertEqual(reopened.active["A2"].value, "cleantext")

    def test_lead_schema_matches_export_columns(self):
        # new_lead() and the export column list must stay in lockstep.
        self.assertEqual(set(lead_crawler.new_lead().keys()), set(lead_crawler.LEAD_COLUMNS))

    def test_codex_manual_requires_seeds_or_fixture(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "out.xlsx"
            args = lead_crawler.parse_args(
                [
                    "--theme", "dpp-rollout-sectors",
                    "--search-provider", "codex_manual",
                    "--output", str(output_path),
                ]
            )
            with self.assertRaises(SystemExit):
                lead_crawler.run(args)

    def test_codex_manual_with_seeds_produces_leads(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            seed_path = Path(temp_dir) / "seeds.txt"
            seed_path.write_text(
                "acme.de\nhttps://www.linkedin.com/company/bravo\n", encoding="utf-8"
            )
            output_path = Path(temp_dir) / "out.xlsx"
            args = lead_crawler.parse_args(
                [
                    "--theme", "linkedin-assisted-cross-reference",
                    "--search-provider", "codex_manual",
                    "--manual-seeds", str(seed_path),
                    "--no-crawl-pages",
                    "--output", str(output_path),
                ]
            )
            lead_crawler.run(args)
            workbook = openpyxl.load_workbook(output_path, data_only=True)
            ws = workbook["Leads"]
            hdr = [c.value for c in ws[1]]
            rows = [dict(zip(hdr, [c.value for c in r])) for r in ws.iter_rows(min_row=2)]
            companies = {r["company_name"] for r in rows}
            # bare domain -> crawlable lead; LinkedIn URL -> reference row.
            self.assertIn("acme.de", companies)
            self.assertIn("Bravo", companies)
            bravo = next(r for r in rows if r["company_name"] == "Bravo")
            self.assertIn("linkedin.com/company/bravo", bravo["linkedin_reference_url"])

    def test_manual_seed_only_theme_requires_seeds(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "out.xlsx"
            args = lead_crawler.parse_args(
                [
                    "--theme", "linkedin-assisted-cross-reference",
                    "--search-provider", "serper",
                    "--search-api-key", "x",
                    "--output", str(output_path),
                ]
            )
            with self.assertRaises(SystemExit):
                lead_crawler.run(args)

    def test_google_cse_requires_cse_id_upfront(self):
        import os as _os
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "out.xlsx"
            args = lead_crawler.parse_args(
                [
                    "--theme", "dpp-rollout-sectors",
                    "--search-provider", "google_cse",
                    "--search-api-key", "x",
                    "--output", str(output_path),
                ]
            )
            env = {k: v for k, v in _os.environ.items() if k != "GOOGLE_CSE_ID"}
            with patch.dict(lead_crawler.os.environ, env, clear=True):
                with self.assertRaises(SystemExit):
                    lead_crawler.run(args)

    def test_run_rejects_malformed_fixture(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            fixture_path = Path(temp_dir) / "broken.json"
            fixture_path.write_text('{"organic_results": [', encoding="utf-8")
            output_path = Path(temp_dir) / "out.xlsx"
            args = lead_crawler.parse_args(
                [
                    "--theme", "dpp-rollout-sectors",
                    "--fixture", str(fixture_path),
                    "--no-crawl-pages",
                    "--output", str(output_path),
                ]
            )
            with self.assertRaises(SystemExit):
                lead_crawler.run(args)

    def test_enrich_public_pages_redacts_key_in_error_notes(self):
        secret_key = "sk_extract_secret_98765"
        theme = lead_crawler.prebuilt_themes()["dpp-rollout-sectors"]
        leads = [{"company_name": "Acme", "domain": "acme.com", "website": "https://acme.com"}]

        def boom(url, provider_id, api_key):
            raise RuntimeError(f"auth failed for header Bearer {secret_key}")

        with patch.object(lead_crawler, "extract_page", side_effect=boom):
            lead_crawler.enrich_public_pages(leads, theme, "firecrawl", secret_key)

        notes = leads[0].get("notes", "")
        self.assertNotIn(secret_key, notes)
        self.assertIn("***", notes)


if __name__ == "__main__":
    unittest.main()
