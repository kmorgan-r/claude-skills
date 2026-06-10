#!/usr/bin/env python3
"""Find company-level cold leads from public web search signals."""

from __future__ import annotations

import argparse
import getpass
import ipaddress
import json
import logging
import os
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import quote, urljoin, urlparse

import openpyxl
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_OUTPUT = "outputs/cold_leads.xlsx"
CONTACT_HINTS = (
    "sustainability",
    "team",
    "people",
    "leadership",
    "management",
    "about",
    "contact",
    "impressum",
    "legal",
)
SECOND_HOP_PAGE_LIMIT = 5
def theme_target_personas(theme: dict[str, Any]) -> str:
    return theme.get("target_personas", "Decision Maker / Department Head / Buyer")


def theme_contact_search_titles(theme: dict[str, Any]) -> tuple[str, ...]:
    return tuple(
        theme.get(
            "contact_search_titles",
            [
                "CEO",
                "VP",
                "Director",
                "Manager",
                "Head of Department",
                "Chief Operating Officer",
                "Founder",
                "Owner",
            ],
        )
    )


def theme_buyer_title_terms(theme: dict[str, Any]) -> tuple[str, ...]:
    return tuple(
        theme.get(
            "buyer_title_terms",
            [
                "manager",
                "director",
                "vp",
                "head",
                "chief",
                "owner",
                "founder",
                "president",
                "partner",
                "principal",
            ],
        )
    )


def theme_lead_signals(theme: dict[str, Any]) -> list[str]:
    return theme.get("lead_signals", [])


def theme_high_priority_title_terms(theme: dict[str, Any]) -> tuple[str, ...]:
    return tuple(theme.get("high_priority_title_terms", []))


def theme_medium_priority_title_terms(theme: dict[str, Any]) -> tuple[str, ...]:
    return tuple(theme.get("medium_priority_title_terms", []))


ROLE_TITLE_TERMS = (
    "advisor",
    "analyst",
    "chief",
    "consultant",
    "coordinator",
    "director",
    "head",
    "lead",
    "manager",
    "officer",
    "partner",
    "president",
    "principal",
    "specialist",
    "vp",
)
BLOCKED_DOMAINS = {
    # people-data vendors / contact databases
    "zoominfo.com",
    "lusha.com",
    "apollo.io",
    "rocketreach.co",
    "ensun.io",
    "cognism.com",
    "seamless.ai",
    "leadiq.com",
    "uplead.com",
    "adapt.io",
    "signalhire.com",
    "contactout.com",
    "snov.io",
    "hunter.io",
    # company directories / listicles / b2b portals
    "kompass.com",
    "europages.com",
    "europages.co.uk",
    "crunchbase.com",
    "dnb.com",
    "yellowpages.com",
    "yelp.com",
    "glassdoor.com",
    "indeed.com",
    "wlw.de",
    "thomasnet.com",
    "tradeindia.com",
    "alibaba.com",
    # social / encyclopedias
    "linkedin.com",
    "facebook.com",
    "twitter.com",
    "x.com",
    "instagram.com",
    "youtube.com",
    "wikipedia.org",
    "wikidata.org",
    "reddit.com",
    "pinterest.com",
    "tiktok.com",
    "medium.com",
    # generic news (intent should come from the company, not press coverage)
    "bloomberg.com",
    "reuters.com",
    "forbes.com",
    "businesswire.com",
    "prnewswire.com",
    "globenewswire.com",
}
_ENTITY_SUFFIX = re.compile(
    r"\b(gmbh|ag|kg|ug|se|ltd|limited|llc|inc|corp|plc|bv|nv|sa|spa|srl|"
    r"oy|ab|as|aps|sas|sarl|pty)\b",
    re.IGNORECASE,
)
EMAIL_RE = re.compile(r"(?<![\w.+-])[\w.+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}(?![\w+-])")
PLACEHOLDER_EMAILS = {
    "email@newsletter.com",
    "you@company.com",
    "user@gmail.com",
    "name@example.com",
    "user@example.com",
    "test@example.com",
}
PLACEHOLDER_LOCAL_PARTS = {"email", "example", "name", "test", "user", "you", "yourname"}
PLACEHOLDER_DOMAINS = {
    "company.com",
    "example.com",
    "newsletter.com",
}
CONSUMER_EMAIL_DOMAINS = {"gmail.com", "hotmail.com", "outlook.com", "yahoo.com"}
_FORMULA_CHARS = ("=", "+", "-", "@")
_BLOCKED_HOSTNAMES = frozenset({
    "localhost",
    "localhost.localdomain",
    "ip6-localhost",
    "ip6-loopback",
    "metadata",
    "metadata.google.internal",
    "metadata.google",
})


@dataclass
class SearchSource:
    query: str
    result_count: int
    source: str


def provider_catalog() -> dict[str, dict[str, dict[str, Any]]]:
    return {
        "search": {
            "serper": {
                "label": "Serper.dev Google Search API",
                "env": "SERPER_API_KEY",
                "requires_key": True,
                "supports_inline_enrichment": True,
            },
            "serpapi": {
                "label": "SerpApi Google Search API",
                "env": "SERPAPI_KEY",
                "requires_key": True,
                "supports_inline_enrichment": True,
            },
            "searchapi": {
                "label": "SearchApi.io Google Search API",
                "env": "SEARCHAPI_API_KEY",
                "requires_key": True,
                "supports_inline_enrichment": True,
            },
            "brave": {
                "label": "Brave Search API",
                "env": "BRAVE_SEARCH_API_KEY",
                "requires_key": True,
                "supports_inline_enrichment": True,
            },
            "tavily": {
                "label": "Tavily Search API",
                "env": "TAVILY_API_KEY",
                "requires_key": True,
                "supports_inline_enrichment": True,
            },
            "exa": {
                "label": "Exa Search API",
                "env": "EXA_API_KEY",
                "requires_key": True,
                "supports_inline_enrichment": True,
            },
            "google_cse": {
                "label": "Google Custom Search JSON API",
                "env": "GOOGLE_API_KEY",
                "requires_key": True,
                "supports_inline_enrichment": True,
                "extra_env": "GOOGLE_CSE_ID",
            },
            "codex_manual": {
                "label": "Codex/manual web research",
                "env": None,
                "requires_key": False,
                "supports_inline_enrichment": True,
            },
        },
        "extract": {
            "codex_builtin": {
                "label": "Built-in requests + BeautifulSoup extraction",
                "env": None,
                "requires_key": False,
            },
            "jina": {
                "label": "Jina Reader",
                "env": "JINA_API_KEY",
                "requires_key": False,
            },
            "firecrawl": {
                "label": "Firecrawl Scrape API",
                "env": "FIRECRAWL_API_KEY",
                "requires_key": True,
            },
            "tavily": {
                "label": "Tavily Extract API",
                "env": "TAVILY_API_KEY",
                "requires_key": True,
            },
            "exa": {
                "label": "Exa Contents API",
                "env": "EXA_API_KEY",
                "requires_key": True,
            },
            "none": {
                "label": "Do not extract page content",
                "env": None,
                "requires_key": False,
            },
        },
    }


def resolve_provider_key(
    provider_id: str,
    provider: dict[str, Any],
    explicit_key: str | None,
    env: dict[str, str] | None = None,
    prompt_fn: Any | None = None,
) -> str | None:
    if explicit_key:
        return explicit_key
    env = env or os.environ
    env_name = provider.get("env")
    if env_name and env.get(env_name):
        return env[env_name]
    if not provider.get("requires_key"):
        return None
    if prompt_fn:
        return prompt_fn(f"Enter API key for {provider_id} ({env_name}): ")
    return None


def prebuilt_themes() -> dict[str, dict[str, Any]]:
    return {
        "generic-b2b": {
            "label": "Generic B2B",
            "description": "A neutral starting point for any B2B lead search. Define your own sectors and keywords via custom theme or interview.",
            "sectors": ["company", "business"],
            "keywords": [],
            "subthemes": ["Custom discovery"],
        },
        "dpp-rollout-sectors": {
            "label": "DPP rollout sectors",
            "description": "EU product sectors facing Digital Product Passport pressure.",
            "sectors": ["textile", "apparel", "footwear", "furniture", "mattress", "toy"],
            "keywords": [
                "manufacturer",
                "supplier",
                "product carbon footprint",
                "sustainability report",
                "Digital Product Passport",
                "PEFCR",
            ],
            "subthemes": ["Textiles / apparel", "Furniture", "Mattresses", "Toys"],
            "target_personas": "Sustainability / ESG / Product Compliance / LCA",
            "contact_search_titles": [
                "Head of Sustainability",
                "ESG Manager",
                "Sustainability Director",
                "Sustainability Manager",
                "ESG Director",
                "Environmental Manager",
                "Product Compliance Manager",
                "Quality Manager",
                "Supply Chain Sustainability Manager",
                "LCA Specialist",
            ],
            "buyer_title_terms": [
                "sustainability",
                "esg",
                "environment",
                "environmental",
                "compliance",
                "product compliance",
                "quality",
                "supply chain",
                "procurement",
                "lca",
                "carbon",
            ],
            "lead_signals": [
                "lca",
                "life-cycle",
                "life cycle",
                "product carbon footprint",
                "iso 14067",
                "pefcr",
                "digital product passport",
                "third-party",
                "sustainability",
            ],
            "high_priority_title_terms": [
                "sustainability",
                "esg",
                "environment",
                "lca",
                "carbon",
            ],
            "medium_priority_title_terms": [
                "compliance",
                "quality",
                "supply chain",
                "procurement",
            ],
        },
        "eu-taxonomy-lca": {
            "label": "EU taxonomy LCA requirements",
            "description": "Activities with EU Taxonomy LCA or life-cycle GHG requirements.",
            "sectors": [
                "hydrogen",
                "chlorine",
                "organic basic chemicals",
                "plastics primary form",
                "hydropower",
                "geothermal energy",
                "renewable fuels",
                "nuclear energy",
                "data-driven GHG reductions",
                "direct air capture",
            ],
            "keywords": [
                "life-cycle GHG",
                "ISO 14067",
                "ISO 14064-1",
                "Commission Recommendation 2013/179/EU",
                "third-party verification",
            ],
            "subthemes": [
                "Product / manufacturing LCA",
                "Energy life-cycle GHG threshold",
                "Digital / ICT avoided-emissions LCA",
                "R&D life-cycle performance evaluation",
                "Adaptation Annex LCA-style requirement",
            ],
            "target_personas": "Sustainability / ESG / Product Compliance / LCA",
            "contact_search_titles": [
                "Head of Sustainability",
                "ESG Manager",
                "Sustainability Director",
                "Sustainability Manager",
                "ESG Director",
                "Environmental Manager",
                "Product Compliance Manager",
                "Quality Manager",
                "Supply Chain Sustainability Manager",
                "LCA Specialist",
            ],
            "buyer_title_terms": [
                "sustainability",
                "esg",
                "environment",
                "environmental",
                "compliance",
                "product compliance",
                "quality",
                "supply chain",
                "procurement",
                "lca",
                "carbon",
            ],
            "lead_signals": [
                "lca",
                "life-cycle",
                "life cycle",
                "product carbon footprint",
                "iso 14067",
                "pefcr",
                "digital product passport",
                "third-party",
                "sustainability",
            ],
            "high_priority_title_terms": [
                "sustainability",
                "esg",
                "environment",
                "lca",
                "carbon",
            ],
            "medium_priority_title_terms": [
                "compliance",
                "quality",
                "supply chain",
                "procurement",
            ],
        },
        "standards-triggered-prospects": {
            "label": "Standards-triggered prospects",
            "description": "Companies publicly mentioning LCA, PCF, PEF, or verification standards.",
            "sectors": ["manufacturer", "supplier", "industrial company", "brand"],
            "keywords": [
                "ISO 14067",
                "ISO 14064-1",
                "PEFCR",
                "product environmental footprint",
                "product carbon footprint",
                "third-party verified",
            ],
            "subthemes": ["ISO 14067", "ISO 14064-1", "PEF / PEFCR", "PCF verification"],
            "target_personas": "Sustainability / ESG / Product Compliance / LCA",
            "contact_search_titles": [
                "Head of Sustainability",
                "ESG Manager",
                "Sustainability Director",
                "Sustainability Manager",
                "ESG Director",
                "Environmental Manager",
                "Product Compliance Manager",
                "Quality Manager",
                "Supply Chain Sustainability Manager",
                "LCA Specialist",
            ],
            "buyer_title_terms": [
                "sustainability",
                "esg",
                "environment",
                "environmental",
                "compliance",
                "product compliance",
                "quality",
                "supply chain",
                "procurement",
                "lca",
                "carbon",
            ],
            "lead_signals": [
                "lca",
                "life-cycle",
                "life cycle",
                "product carbon footprint",
                "iso 14067",
                "pefcr",
                "digital product passport",
                "third-party",
                "sustainability",
            ],
            "high_priority_title_terms": [
                "sustainability",
                "esg",
                "environment",
                "lca",
                "carbon",
            ],
            "medium_priority_title_terms": [
                "compliance",
                "quality",
                "supply chain",
                "procurement",
            ],
        },
        "linkedin-assisted-cross-reference": {
            "label": "LinkedIn-assisted cross-reference",
            "description": "Use user-provided LinkedIn URLs or names as manual seeds, then verify on the public web.",
            "sectors": ["manufacturer", "brand", "supplier"],
            "keywords": [
                "contact",
                "sustainability",
                "environment",
                "product carbon footprint",
                "LCA",
            ],
            "subthemes": ["Manual LinkedIn company URLs", "Manual person/company seeds"],
            "manual_seed_only": True,
            "guardrail": "Do not crawl LinkedIn or automate LinkedIn sessions. Use only user-provided or licensed LinkedIn references.",
            "target_personas": "Sustainability / ESG / Product Compliance / LCA",
            "contact_search_titles": [
                "Head of Sustainability",
                "ESG Manager",
                "Sustainability Director",
                "Sustainability Manager",
                "ESG Director",
                "Environmental Manager",
                "Product Compliance Manager",
                "Quality Manager",
                "Supply Chain Sustainability Manager",
                "LCA Specialist",
            ],
            "buyer_title_terms": [
                "sustainability",
                "esg",
                "environment",
                "environmental",
                "compliance",
                "product compliance",
                "quality",
                "supply chain",
                "procurement",
                "lca",
                "carbon",
            ],
            "lead_signals": [
                "lca",
                "life-cycle",
                "life cycle",
                "product carbon footprint",
                "iso 14067",
                "pefcr",
                "digital product passport",
                "third-party",
                "sustainability",
            ],
            "high_priority_title_terms": [
                "sustainability",
                "esg",
                "environment",
                "lca",
                "carbon",
            ],
            "medium_priority_title_terms": [
                "compliance",
                "quality",
                "supply chain",
                "procurement",
            ],
        },
    }


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Find cold leads and export an Excel workbook.")
    catalog = provider_catalog()
    parser.add_argument("--theme", choices=sorted(prebuilt_themes()), required=False)
    parser.add_argument("--custom-theme-file", help="JSON file with sectors, keywords, and optional subthemes.")
    parser.add_argument("--location", default="European Union", help="Search location phrase to include in queries.")
    parser.add_argument("--max-queries", type=int, default=12)
    parser.add_argument("--max-results", type=int, default=40)
    parser.add_argument("--search-provider", choices=sorted(catalog["search"]), default="serper")
    parser.add_argument("--extract-provider", choices=sorted(catalog["extract"]), default="codex_builtin")
    parser.add_argument("--search-api-key", help="API key for the selected search provider.")
    parser.add_argument("--extract-api-key", help="API key for the selected extract provider.")
    parser.add_argument("--serpapi-key", default=os.environ.get("SERPAPI_KEY"), help="Deprecated; use --search-api-key.")
    parser.add_argument("--prompt-for-keys", action="store_true", help="Prompt securely for missing required API keys.")
    parser.add_argument("--fixture", help="Read SerpApi-style JSON from a local fixture instead of calling SerpApi.")
    parser.add_argument("--manual-seeds", help="CSV/JSON/TXT file of user-provided companies, domains, or LinkedIn URLs.")
    parser.add_argument("--output", default=DEFAULT_OUTPUT)
    parser.add_argument("--no-crawl-pages", action="store_true", help="Do not fetch candidate company pages.")
    parser.add_argument("--contact-search", action="store_true", help="Run targeted person/title searches for each candidate company.")
    parser.add_argument("--contact-search-queries", type=int, default=4, help="Max contact-search queries per lead.")
    parser.add_argument("--contact-search-budget", type=int, default=0, help="Max total contact-search queries across all leads (0 = unlimited).")
    parser.add_argument("--list-themes", action="store_true")
    parser.add_argument("--list-providers", action="store_true")
    return parser.parse_args(argv)


def load_theme(args: argparse.Namespace) -> tuple[str, dict[str, Any]]:
    if args.custom_theme_file:
        data = json.loads(Path(args.custom_theme_file).read_text(encoding="utf-8"))
        return data.get("id", "custom"), data
    if args.theme:
        return args.theme, prebuilt_themes()[args.theme]
    raise SystemExit("Choose --theme, pass --custom-theme-file, or use --list-themes.")


def expand_queries(theme: dict[str, Any], location: str, max_queries: int) -> list[str]:
    sectors = theme.get("sectors") or ["company"]
    keywords = theme.get("keywords") or []
    queries: list[str] = []
    for sector in sectors:
        if keywords:
            for keyword in keywords:
                queries.append(f'"{sector}" "{keyword}" {location}')
        else:
            queries.append(f'"{sector}" {location}')
    return queries[: max(1, max_queries)]


MULTI_LABEL_SUFFIXES = {
    "co.uk", "org.uk", "gov.uk", "ac.uk", "ltd.uk", "plc.uk", "me.uk",
    "com.au", "net.au", "org.au", "co.nz", "co.za", "co.jp", "or.jp",
    "com.br", "com.mx", "com.tr", "com.cn", "com.sg", "com.hk", "co.in",
    "com.es", "com.pl", "com.ua",
}


def normalized_domain(url: str) -> str:
    # .hostname (not .netloc) so explicit ports and userinfo never leak into
    # blocklist or dedup comparisons.
    host = (urlparse(url).hostname or "").lower()
    if host.startswith("www."):
        host = host[4:]
    return host


def registrable_domain(url: str) -> str:
    """Return the eTLD+1 (registrable domain) for a URL, lowercased."""
    host = normalized_domain(url)
    if not host or "." not in host:
        return host
    parts = host.split(".")
    last2 = ".".join(parts[-2:])
    if len(parts) >= 3 and last2 in MULTI_LABEL_SUFFIXES:
        return ".".join(parts[-3:])
    return last2


def _is_private_ip_url(url: str) -> bool:
    host = urlparse(url).hostname or ""
    host_lower = host.lower()
    if host_lower in _BLOCKED_HOSTNAMES:
        return True
    if host_lower.endswith((".local", ".internal")):
        return True
    try:
        ip = ipaddress.ip_address(host)
        return ip.is_private or ip.is_loopback or ip.is_link_local or ip.is_reserved
    except ValueError:
        return False


def is_blocked_url(url: str) -> bool:
    domain = normalized_domain(url)
    return any(domain == blocked or domain.endswith(f".{blocked}") for blocked in BLOCKED_DOMAINS)


def leads_from_search_results(results: list[dict[str, Any]], query: str, theme_id: str, theme: dict[str, Any]) -> list[dict[str, Any]]:
    seen: set[str] = set()
    leads: list[dict[str, Any]] = []
    for item in results:
        url = item.get("link") or item.get("url") or ""
        if not url or is_blocked_url(url) or _is_private_ip_url(url):
            continue
        domain = normalized_domain(url)
        company_key = registrable_domain(url)
        if not domain or company_key in seen:
            continue
        seen.add(company_key)
        title = item.get("title") or domain
        snippet = item.get("snippet") or ""
        company_name = clean_company_name(title, domain)
        if looks_like_bad_company_name(company_name):
            continue
        leads.append(
            {
                "company_name": company_name,
                "domain": domain,
                "website": f"{urlparse(url).scheme or 'https'}://{urlparse(url).netloc}",
                "country": "",
                "sector": "",
                "theme": theme_id,
                "matched_signal": snippet[:500],
                "target_persona": theme_target_personas(theme),
                "contact_name": "",
                "contact_title": "",
                "contact_email": "",
                "contact_page": "",
                "contact_link": "",
                "contact_source_url": "",
                "contact_confidence": "",
                "contact_data_type": "company",
                "person_source_type": "",
                "public_profile_url": "",
                "email_discovery_method": "",
                "email_verification_status": "",
                "email_confidence": "",
                "do_not_contact_reason": "",
                "linkedin_reference_url": "",
                "lead_score": score_lead(snippet, theme),
                "source_url": url,
                "evidence_snippet": snippet,
                "business_relevance_basis": query,
                "consent_status": "unknown",
                "outreach_allowed_review": "needs review",
                "delete_if_not_used_by": "",
                "notes": "",
                "odoo_ready": "no",
            }
        )
    return leads


def clean_company_name(title: str, domain: str) -> str:
    title = re.split(r"\s+[-|]\s+", title.strip())[0]
    title = re.sub(r"\s+", " ", title)
    return title or domain


def looks_like_bad_company_name(name: str) -> bool:
    """Heuristic guard: True if `name` looks like a SERP title, not a company."""
    if not name or not name.strip():
        return True
    n = name.strip()
    if "…" in n or "..." in n:
        return True
    if len(n) > 80:
        return True
    if ":" in n or " | " in n or " — " in n or " - " in n:
        return True
    words = n.split()
    if len(words) > 8:
        return True
    # Listicle / directory titles ("Top 100 ... Companies in Germany (2026)").
    if re.search(r"\b(companies|list|ranking|directory|top\s+\d+|best\s+\d+)\b", n, re.IGNORECASE):
        return True
    if re.search(r"\(?(19|20)\d{2}\)?", n):
        return True
    # A legal suffix that is NOT at/near the end => the name is a phrase that
    # happens to contain a company, e.g. "...innovation and Storchenwiege GmbH".
    matches = list(_ENTITY_SUFFIX.finditer(n))
    if matches:
        last = matches[-1]
        tail = n[last.end():].strip(" .,&")
        tail_words = [w for w in tail.split() if w and not _ENTITY_SUFFIX.fullmatch(w.strip(".,&"))]
        if len(tail_words) > 2:  # allow short tails like "& Co. KG"
            return True
    return False


def score_lead(text: str, theme: dict[str, Any]) -> int:
    lowered = text.lower()
    signals = theme_lead_signals(theme)
    if signals:
        return min(100, 30 + sum(10 for signal in signals if signal in lowered))
    return 30


def contact_search_queries(lead: dict[str, Any], theme: dict[str, Any], max_queries: int = 12) -> list[str]:
    company_name = lead.get("company_name") or lead.get("domain") or ""
    domain = normalized_domain(lead.get("website") or lead.get("domain") or "")
    titles = theme_contact_search_titles(theme)
    queries: list[str] = []
    for title in titles:
        if company_name:
            queries.append(f'"{company_name}" "{title}"')
        if domain and len(queries) < max_queries:
            queries.append(f"site:{domain} \"{title}\"")
    if domain and titles and not any(query.startswith(f"site:{domain} ") for query in queries[:max_queries]):
        queries.insert(1 if queries else 0, f"site:{domain} \"{titles[0]}\"")
    for title in titles:
        if domain and f"site:{domain} \"{title}\"" not in queries:
            queries.append(f"site:{domain} \"{title}\"")
    return queries[: max(1, max_queries)]


def enrich_contacts_via_search(
    leads: list[dict[str, Any]],
    provider_id: str,
    api_key: str,
    theme: dict[str, Any],
    per_lead_queries: int = 4,
    per_query_results: int = 3,
    budget: int = 0,
) -> list[SearchSource]:
    eligible_count = sum(1 for lead in leads if lead.get("contact_data_type") != "person")
    estimated = eligible_count * per_lead_queries
    if budget > 0:
        estimated = min(estimated, budget)
        print(f"Estimated contact-search queries: {estimated} (budget: {budget})")
    else:
        print(f"Estimated contact-search queries: {estimated} (budget: unlimited)")
    sources: list[SearchSource] = []
    queries_used = 0
    for lead in leads:
        try:
            if lead.get("contact_data_type") == "person":
                continue
            for query in contact_search_queries(lead, theme, per_lead_queries):
                if budget > 0 and queries_used >= budget:
                    print(f"Contact-search budget exhausted ({budget}); stopping.")
                    return sources
                queries_used += 1
                try:
                    results = search_provider(query, provider_id, api_key, per_query_results)
                except Exception as exc:  # noqa: BLE001 - keep other leads available
                    safe_exc = _redact_key(str(exc), api_key)
                    lead["notes"] = append_note(lead.get("notes", ""), f"Contact search failed ({query}): {safe_exc}")
                    continue
                sources.append(SearchSource(query, len(results), f"{provider_id}:contact_search"))
                apply_contact_search_results(lead, results, query, theme)
                if lead.get("contact_data_type") == "person":
                    break
        except Exception as exc:  # noqa: BLE001 - flush partial work on unexpected failure
            safe_exc = _redact_key(str(exc), api_key)
            logging.warning("Unexpected error during contact search for %s: %s", lead.get("company_name", "unknown"), safe_exc)
            lead["notes"] = append_note(lead.get("notes", ""), f"Contact search interrupted: {safe_exc}")
            continue
    return sources


def apply_contact_search_results(lead: dict[str, Any], results: list[dict[str, Any]], query: str, theme: dict[str, Any]) -> None:
    for item in results:
        url = item.get("link") or item.get("url") or ""
        if not url or is_blocked_url(url):
            continue
        title = item.get("title") or ""
        snippet = item.get("snippet") or item.get("content") or item.get("text") or ""
        people = extract_contact_people(f"{title}\n{snippet}", url, theme)
        if not people:
            continue
        person = people[0]
        if should_replace_person(lead, person):
            lead["contact_name"] = person["contact_name"]
            lead["contact_title"] = person["contact_title"]
            lead["contact_email"] = person.get("contact_email") or lead.get("contact_email", "")
            lead["contact_link"] = person.get("contact_link") or url
            lead["contact_source_url"] = url
            lead["contact_confidence"] = max(0, int(person.get("contact_confidence") or 0) - 10)
            lead["contact_data_type"] = "person"
            lead["person_source_type"] = classify_person_source(url)
            lead["public_profile_url"] = url
            lead["email_discovery_method"] = "public_snippet" if person.get("contact_email") else "none"
            lead["email_verification_status"] = "unverified" if person.get("contact_email") else ""
            lead["email_confidence"] = 40 if person.get("contact_email") else ""
            lead["business_relevance_basis"] = append_note(lead.get("business_relevance_basis", ""), query)
            return


def classify_person_source(url: str) -> str:
    domain = normalized_domain(url)
    path = urlparse(url).path.lower()
    if "linkedin.com" in domain:
        return "linkedin_reference"
    if any(part in path for part in ("event", "speaker", "conference", "webinar")):
        return "event_or_speaker_page"
    if path.endswith(".pdf"):
        return "pdf_or_report"
    return "public_web"


def _raise_for_status_redacted(response: requests.Response, api_key: str) -> None:
    """raise_for_status, but scrub the API key from the error message.

    requests.HTTPError includes the full request URL, which leaks keys passed
    as query params (SerpApi api_key=, SearchApi api_key=, Google CSE key=).
    """
    try:
        response.raise_for_status()
    except requests.HTTPError as exc:
        raise requests.HTTPError(
            _redact_key(str(exc), api_key), response=response
        ) from None


def serpapi_search(query: str, api_key: str, max_results: int) -> list[dict[str, Any]]:
    collected: list[dict[str, Any]] = []
    start = 0
    while len(collected) < max_results:
        response = requests.get(
            "https://serpapi.com/search.json",
            params={
                "engine": "google",
                "q": query,
                "api_key": api_key,
                "num": min(10, max_results - len(collected)),
                "start": start,
            },
            timeout=30,
        )
        _raise_for_status_redacted(response, api_key)
        payload = response.json()
        page_results = payload.get("organic_results") or []
        if not page_results:
            break
        collected.extend(page_results)
        start += len(page_results)
        if not payload.get("serpapi_pagination", {}).get("next"):
            break
        time.sleep(1)
    return collected[:max_results]


def search_provider(query: str, provider_id: str, api_key: str, max_results: int) -> list[dict[str, Any]]:
    if provider_id == "serpapi":
        return serpapi_search(query, api_key, max_results)
    if provider_id == "serper":
        return serper_search(query, api_key, max_results)
    if provider_id == "searchapi":
        return searchapi_search(query, api_key, max_results)
    if provider_id == "brave":
        return brave_search(query, api_key, max_results)
    if provider_id == "tavily":
        return tavily_search(query, api_key, max_results)
    if provider_id == "exa":
        return exa_search(query, api_key, max_results)
    if provider_id == "google_cse":
        return google_cse_search(query, api_key, max_results)
    raise SystemExit(f"Search provider {provider_id!r} is not runnable from the script. Use manual seeds or fixture.")


def serper_search(query: str, api_key: str, max_results: int) -> list[dict[str, Any]]:
    response = requests.post(
        "https://google.serper.dev/search",
        headers={"X-API-KEY": api_key, "Content-Type": "application/json"},
        json={"q": query, "num": min(max_results, 100)},
        timeout=30,
    )
    response.raise_for_status()
    payload = response.json()
    return payload.get("organic") or []


def searchapi_search(query: str, api_key: str, max_results: int) -> list[dict[str, Any]]:
    response = requests.get(
        "https://www.searchapi.io/api/v1/search",
        params={"engine": "google", "q": query, "api_key": api_key, "num": min(max_results, 100)},
        timeout=30,
    )
    _raise_for_status_redacted(response, api_key)
    payload = response.json()
    return payload.get("organic_results") or []


def brave_search(query: str, api_key: str, max_results: int) -> list[dict[str, Any]]:
    response = requests.get(
        "https://api.search.brave.com/res/v1/web/search",
        headers={"X-Subscription-Token": api_key, "Accept": "application/json"},
        params={"q": query, "count": min(max_results, 20)},
        timeout=30,
    )
    response.raise_for_status()
    payload = response.json()
    return [
        {
            "title": item.get("title"),
            "link": item.get("url"),
            "snippet": item.get("description"),
        }
        for item in payload.get("web", {}).get("results", [])
    ]


def tavily_search(query: str, api_key: str, max_results: int) -> list[dict[str, Any]]:
    response = requests.post(
        "https://api.tavily.com/search",
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json={"query": query, "max_results": min(max_results, 20), "include_answer": False},
        timeout=30,
    )
    response.raise_for_status()
    payload = response.json()
    return [
        {
            "title": item.get("title"),
            "link": item.get("url"),
            "snippet": item.get("content"),
        }
        for item in payload.get("results", [])
    ]


def exa_search(query: str, api_key: str, max_results: int) -> list[dict[str, Any]]:
    response = requests.post(
        "https://api.exa.ai/search",
        headers={"x-api-key": api_key, "Content-Type": "application/json"},
        json={"query": query, "numResults": min(max_results, 100), "contents": {"text": True}},
        timeout=30,
    )
    response.raise_for_status()
    payload = response.json()
    return [
        {
            "title": item.get("title"),
            "link": item.get("url"),
            "snippet": item.get("text") or item.get("summary"),
        }
        for item in payload.get("results", [])
    ]


def google_cse_search(query: str, api_key: str, max_results: int) -> list[dict[str, Any]]:
    cse_id = os.environ.get("GOOGLE_CSE_ID")
    if not cse_id:
        raise SystemExit("Set GOOGLE_CSE_ID when using --search-provider google_cse.")
    response = requests.get(
        "https://www.googleapis.com/customsearch/v1",
        params={"key": api_key, "cx": cse_id, "q": query, "num": min(max_results, 10)},
        timeout=30,
    )
    _raise_for_status_redacted(response, api_key)
    payload = response.json()
    return payload.get("items") or []


def read_fixture(path: str) -> list[dict[str, Any]]:
    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    if isinstance(payload, list):
        return payload
    return payload.get("organic_results") or payload.get("organic") or payload.get("results") or []


def read_manual_seeds(path: str) -> list[dict[str, Any]]:
    source = Path(path)
    text = source.read_text(encoding="utf-8")
    if source.suffix.lower() == ".json":
        raw = json.loads(text)
        entries = raw if isinstance(raw, list) else raw.get("seeds", [])
    else:
        entries = [line.strip() for line in text.splitlines() if line.strip()]

    results: list[dict[str, Any]] = []
    for entry in entries:
        if isinstance(entry, dict):
            name = entry.get("company") or entry.get("name") or entry.get("url") or "Manual seed"
            url = entry.get("website") or entry.get("url") or ""
            linkedin = entry.get("linkedin") or ""
        else:
            name = str(entry)
            url = str(entry) if str(entry).startswith(("http://", "https://")) else ""
            linkedin = url if "linkedin.com/" in url else ""
        if "linkedin.com/" in url:
            url = ""
        if not url:
            logging.warning("Skipping manual seed %r: no usable URL", name)
            continue
        results.append(
            {
                "title": name,
                "link": url,
                "snippet": "Manual seed supplied by user; verify contact evidence on public non-LinkedIn sources.",
                "linkedin_reference_url": linkedin,
            }
        )
    return results


def enrich_public_pages(leads: list[dict[str, Any]], theme: dict[str, Any], provider_id: str = "codex_builtin", api_key: str | None = None) -> None:
    if provider_id == "none":
        return
    for lead in leads:
        try:
            page = extract_page(lead["website"], provider_id, api_key)
            apply_page_enrichment(lead, page, theme)
            html = page.get("html") or page.get("text") or ""
            second_hop_urls = candidate_contact_links(html, lead["website"], SECOND_HOP_PAGE_LIMIT)
            visited = {lead["website"].rstrip("/")}
            for url in second_hop_urls:
                normalized_url = url.rstrip("/")
                if normalized_url in visited:
                    continue
                visited.add(normalized_url)
                try:
                    linked_page = extract_page(url, provider_id, api_key)
                except Exception as exc:  # noqa: BLE001 - keep other pages available
                    lead["notes"] = append_note(lead.get("notes", ""), f"Linked page crawl failed ({url}): {exc}")
                    continue
                apply_page_enrichment(lead, linked_page, theme)
                if lead.get("contact_data_type") == "person" and lead.get("contact_email"):
                    break
        except Exception as exc:  # noqa: BLE001 - enrichment should not stop export
            lead["notes"] = append_note(lead.get("notes", ""), f"Page crawl failed: {exc}")


def _redact_key(text: str, key: str) -> str:
    if key and len(key) > 3:
        return text.replace(key, "***")
    return text


def append_note(existing: str | None, note: str) -> str:
    if not existing:
        return note
    return f"{existing}; {note}"


def apply_page_enrichment(lead: dict[str, Any], page: dict[str, Any], theme: dict[str, Any]) -> None:
    html = page.get("html") or page.get("text") or ""
    page_url = page.get("url") or lead["website"]
    emails = page.get("emails") or public_emails(html)
    if emails and not lead.get("contact_email"):
        lead["contact_email"] = emails[0]
    contact = page.get("contact_page") or find_contact_link(html, page_url)
    if contact and not lead.get("contact_page"):
        lead["contact_page"] = contact
    people = extract_contact_people(html, page_url, theme)
    if people and should_replace_person(lead, people[0]):
        person = people[0]
        lead["contact_name"] = person["contact_name"]
        lead["contact_title"] = person["contact_title"]
        lead["contact_email"] = person.get("contact_email") or lead.get("contact_email", "")
        lead["contact_link"] = person.get("contact_link", "")
        lead["contact_source_url"] = person["contact_source_url"]
        lead["contact_confidence"] = person["contact_confidence"]
        lead["contact_data_type"] = "person"
    if page.get("text") and not lead.get("evidence_snippet"):
        lead["evidence_snippet"] = page["text"][:500]


def should_replace_person(lead: dict[str, Any], candidate: dict[str, Any]) -> bool:
    existing_confidence = lead.get("contact_confidence") or 0
    try:
        existing_confidence = int(existing_confidence)
    except (TypeError, ValueError):
        existing_confidence = 0
    return not lead.get("contact_name") or int(candidate.get("contact_confidence") or 0) > existing_confidence


def extract_page(url: str, provider_id: str, api_key: str | None) -> dict[str, Any]:
    if provider_id == "codex_builtin":
        html = fetch_text(url)
        return {"url": url, "html": html, "text": BeautifulSoup(html, "html.parser").get_text(" ", strip=True)}
    if provider_id == "jina":
        return jina_extract(url, api_key)
    if provider_id == "firecrawl":
        return firecrawl_extract(url, api_key or "")
    if provider_id == "tavily":
        return tavily_extract(url, api_key or "")
    if provider_id == "exa":
        return exa_extract(url, api_key or "")
    raise SystemExit(f"Unknown extract provider: {provider_id}")


def jina_extract(url: str, api_key: str | None) -> dict[str, Any]:
    parsed = urlparse(url)
    if parsed.scheme not in ("http", "https"):
        return {"url": url, "text": "", "emails": []}
    if _is_private_ip_url(url):
        return {"url": url, "text": "", "emails": []}
    encoded = quote(url, safe="")
    headers = {"Authorization": f"Bearer {api_key}"} if api_key else {}
    response = requests.get(f"https://r.jina.ai/{encoded}", headers=headers, timeout=30)
    response.raise_for_status()
    text = response.text
    return {"url": url, "text": text, "emails": public_emails(text)}


def firecrawl_extract(url: str, api_key: str) -> dict[str, Any]:
    parsed = urlparse(url)
    if parsed.scheme not in ("http", "https") or _is_private_ip_url(url):
        return {"url": url, "text": "", "html": "", "emails": []}
    response = requests.post(
        "https://api.firecrawl.dev/v1/scrape",
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json={"url": url, "formats": ["markdown", "html"]},
        timeout=45,
    )
    response.raise_for_status()
    payload = response.json()
    data = payload.get("data") or payload
    text = data.get("markdown") or data.get("content") or ""
    html = data.get("html") or ""
    return {"url": url, "text": text, "html": html, "emails": public_emails(" ".join([text, html]))}


def tavily_extract(url: str, api_key: str) -> dict[str, Any]:
    parsed = urlparse(url)
    if parsed.scheme not in ("http", "https") or _is_private_ip_url(url):
        return {"url": url, "text": "", "emails": []}
    response = requests.post(
        "https://api.tavily.com/extract",
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json={"urls": [url]},
        timeout=45,
    )
    response.raise_for_status()
    payload = response.json()
    item = (payload.get("results") or [{}])[0]
    text = item.get("raw_content") or item.get("content") or ""
    return {"url": url, "text": text, "emails": public_emails(text)}


def exa_extract(url: str, api_key: str) -> dict[str, Any]:
    parsed = urlparse(url)
    if parsed.scheme not in ("http", "https") or _is_private_ip_url(url):
        return {"url": url, "text": "", "emails": []}
    response = requests.post(
        "https://api.exa.ai/contents",
        headers={"x-api-key": api_key, "Content-Type": "application/json"},
        json={"urls": [url], "text": True},
        timeout=45,
    )
    response.raise_for_status()
    payload = response.json()
    item = (payload.get("results") or [{}])[0]
    text = item.get("text") or ""
    return {"url": url, "text": text, "emails": public_emails(text)}


def fetch_text(url: str) -> str:
    parsed = urlparse(url)
    if parsed.scheme not in ("http", "https"):
        return ""
    if _is_private_ip_url(url):
        return ""
    current_url = url
    for _ in range(5):  # max redirect hops
        parsed = urlparse(current_url)
        if parsed.scheme not in ("http", "https"):
            return ""
        if _is_private_ip_url(current_url):
            return ""
        response = requests.get(
            current_url,
            timeout=15,
            allow_redirects=False,
            headers={"User-Agent": "B2B lead research bot; contact via website"},
        )
        response.raise_for_status()
        if not response.is_redirect:
            return response.text[:500_000]
        location = response.headers.get("Location", "")
        current_url = urljoin(current_url, location)
    return ""  # too many redirects


def public_emails(html: str) -> list[str]:
    emails = []
    for email in EMAIL_RE.findall(html):
        lowered = email.lower()
        if lowered.endswith((".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg")):
            continue
        local_part, domain = lowered.split("@", 1)
        if lowered in PLACEHOLDER_EMAILS or local_part in PLACEHOLDER_LOCAL_PARTS:
            continue
        if domain in PLACEHOLDER_DOMAINS or domain in CONSUMER_EMAIL_DOMAINS:
            continue
        if lowered not in emails:
            emails.append(lowered)
    return emails[:5]


def extract_contact_people(html: str, source_url: str, theme: dict[str, Any]) -> list[dict[str, Any]]:
    soup = BeautifulSoup(html, "html.parser")
    text = soup.get_text("\n", strip=True)
    emails = public_emails(html)
    people: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    buyer_terms = theme_buyer_title_terms(theme)

    for line in text.splitlines():
        line = re.sub(r"\s+", " ", line).strip()
        if not line or len(line) > 180:
            continue
        if "****" in line or "******" in line:
            continue
        if not any(term in line.lower() for term in buyer_terms):
            continue
        match = re.search(
            r"\b([A-Z][A-Za-z.'-]+(?:\s+[A-Z][A-Za-z.'-]+){1,3})\s*[,|-]\s*([^|\n,]{3,90})",
            line,
        )
        if not match:
            continue
        name = match.group(1).strip()
        title = match.group(2).strip()
        if not looks_like_person_name(name):
            continue
        if not looks_like_contact_title(title, theme):
            continue
        key = (name.lower(), title.lower())
        if key in seen:
            continue
        seen.add(key)
        people.append(
            {
                "contact_name": name,
                "contact_title": title,
                "contact_email": best_email_for_name(name, emails),
                "contact_link": "",
                "contact_source_url": source_url,
                "contact_confidence": contact_confidence(title, bool(emails), theme),
            }
        )

    return sorted(people, key=lambda item: item["contact_confidence"], reverse=True)


def looks_like_person_name(name: str) -> bool:
    lowered = name.lower()
    bad_terms = {"contact us", "privacy policy", "data protection", "sustainability report"}
    generic_name_terms = {
        "accounting",
        "advisor",
        "analyst",
        "audit",
        "carbon",
        "chief",
        "compliance",
        "consultant",
        "coordinator",
        "corporate",
        "director",
        "environmental",
        "esg",
        "footprint",
        "head",
        "lead",
        "manager",
        "officer",
        "partner",
        "principal",
        "reporting",
        "specialist",
        "sustainability",
    }
    if lowered in bad_terms:
        return False
    parts = name.split()
    if any(part.strip(".,").isupper() and len(part.strip(".,")) > 1 for part in parts):
        return False
    if any(part.lower().strip(".,:;()[]") in generic_name_terms for part in parts):
        return False
    return 2 <= len(parts) <= 4 and all(part[0].isupper() for part in parts if part)


def looks_like_contact_title(title: str, theme: dict[str, Any]) -> bool:
    lowered = title.lower()
    buyer_terms = theme_buyer_title_terms(theme)
    if buyer_terms and not any(term in lowered for term in buyer_terms):
        return False
    if not any(re.search(rf"(^|[^a-z]){re.escape(term)}([^a-z]|$)", lowered) for term in ROLE_TITLE_TERMS):
        return False
    if re.search(r"\b(is|are|was|were|now|available|tracking|performance)\b", lowered):
        return False
    return True


def best_email_for_name(name: str, emails: list[str]) -> str:
    if not emails:
        return ""
    parts = [part.lower() for part in re.split(r"\s+", name) if len(part) > 1]
    for email in emails:
        local = email.split("@", 1)[0].lower()
        if all(part in local for part in (parts[0], parts[-1])):
            return email
    for email in emails:
        local = email.split("@", 1)[0].lower()
        if any(part in local for part in parts):
            return email
    return emails[0]


def contact_confidence(title: str, has_email: bool, theme: dict[str, Any]) -> int:
    score = 55
    lowered = title.lower()
    high = theme_high_priority_title_terms(theme)
    medium = theme_medium_priority_title_terms(theme)
    if high and any(term in lowered for term in high):
        score += 25
    elif medium and any(term in lowered for term in medium):
        score += 15
    if has_email:
        score += 10
    return min(100, score)


def find_contact_link(html: str, base_url: str) -> str:
    links = candidate_contact_links(html, base_url, 1)
    return links[0] if links else ""


def candidate_contact_links(html: str, base_url: str, limit: int = SECOND_HOP_PAGE_LIMIT) -> list[str]:
    soup = BeautifulSoup(html, "html.parser")
    base_domain = normalized_domain(base_url)
    candidates: list[tuple[int, str]] = []
    seen: set[str] = set()
    for anchor in soup.find_all("a", href=True):
        href = anchor["href"]
        url = urljoin(base_url, href)
        parsed = urlparse(url)
        if parsed.scheme not in ("http", "https"):
            continue
        if _is_private_ip_url(url):
            continue
        if normalized_domain(url) != base_domain:
            continue
        if is_blocked_url(url):
            continue
        if re.search(r"\.(pdf|jpg|jpeg|png|gif|webp|svg|zip)(\?|$)", parsed.path.lower()):
            continue
        link_text = anchor.get_text(" ", strip=True).lower()
        path = parsed.path.lower()
        label = f"{link_text} {path}"
        matched_hints = [
            hint
            for hint in CONTACT_HINTS
            if re.search(rf"(^|[^a-z]){re.escape(hint)}([^a-z]|$)", label)
        ]
        if not matched_hints:
            continue
        normalized_url = url.split("#", 1)[0].rstrip("/")
        if normalized_url in seen:
            continue
        seen.add(normalized_url)
        priority = min(CONTACT_HINTS.index(hint) for hint in matched_hints)
        candidates.append((priority, normalized_url))
    return [url for _, url in sorted(candidates, key=lambda item: item[0])[:limit]]


def merge_linkedin_references(leads: list[dict[str, Any]], manual_results: list[dict[str, Any]]) -> None:
    linkedin_by_name = {
        (item.get("title") or "").lower(): item.get("linkedin_reference_url") or ""
        for item in manual_results
        if item.get("linkedin_reference_url")
    }
    for lead in leads:
        key = lead["company_name"].lower()
        if key in linkedin_by_name:
            lead["linkedin_reference_url"] = linkedin_by_name[key]


def export_workbook(
    leads: list[dict[str, Any]],
    sources: list[SearchSource],
    rejected: list[dict[str, Any]],
    config: dict[str, Any],
    output_path: str,
) -> None:
    workbook = openpyxl.Workbook()
    leads_sheet = workbook.active
    leads_sheet.title = "Leads"
    source_sheet = workbook.create_sheet("Sources")
    rejected_sheet = workbook.create_sheet("Rejected")
    config_sheet = workbook.create_sheet("Run Config")

    lead_columns = [
        "company_name",
        "domain",
        "website",
        "country",
        "sector",
        "theme",
        "matched_signal",
        "target_persona",
        "contact_name",
        "contact_title",
        "contact_email",
        "contact_page",
        "contact_link",
        "contact_source_url",
        "contact_confidence",
        "contact_data_type",
        "person_source_type",
        "public_profile_url",
        "email_discovery_method",
        "email_verification_status",
        "email_confidence",
        "do_not_contact_reason",
        "linkedin_reference_url",
        "lead_score",
        "source_url",
        "evidence_snippet",
        "business_relevance_basis",
        "consent_status",
        "outreach_allowed_review",
        "delete_if_not_used_by",
        "notes",
        "odoo_ready",
    ]
    write_table(leads_sheet, lead_columns, [[lead.get(col, "") for col in lead_columns] for lead in leads])

    source_columns = ["query", "result_count", "source"]
    write_table(source_sheet, source_columns, [[s.query, s.result_count, s.source] for s in sources])

    rejected_columns = ["url", "reason"]
    write_table(rejected_sheet, rejected_columns, [[row.get("url", ""), row.get("reason", "")] for row in rejected])

    config_rows = [[key, json.dumps(value) if isinstance(value, (dict, list)) else value] for key, value in config.items()]
    write_table(config_sheet, ["key", "value"], config_rows)

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        autosize(sheet)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def write_table(sheet: openpyxl.worksheet.worksheet.Worksheet, headers: list[str], rows: list[list[Any]]) -> None:
    sheet.append(headers)
    for row in rows:
        sanitized: list[Any] = []
        for v in row:
            if isinstance(v, str) and v.startswith(_FORMULA_CHARS):
                sanitized.append("'" + v)
            else:
                sanitized.append(v)
        sheet.append(sanitized)
    header_fill = PatternFill("solid", fgColor="1F4E5F")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill


def autosize(sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for col_idx, column in enumerate(sheet.iter_cols(), start=1):
        width = min(60, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def list_themes() -> None:
    for theme_id, theme in prebuilt_themes().items():
        print(f"{theme_id}: {theme['label']} - {theme['description']}")


def list_providers() -> None:
    catalog = provider_catalog()
    for category in ("search", "extract"):
        print(f"{category} providers:")
        for provider_id, provider in catalog[category].items():
            env_name = provider.get("env") or "no key"
            print(f"  {provider_id}: {provider['label']} ({env_name})")


def run(args: argparse.Namespace) -> Path:
    if args.list_themes:
        list_themes()
        return Path(args.output)
    if args.list_providers:
        list_providers()
        return Path(args.output)

    theme_id, theme = load_theme(args)
    queries = expand_queries(theme, args.location, args.max_queries)
    sources: list[SearchSource] = []
    rejected: list[dict[str, Any]] = []
    raw_results: list[dict[str, Any]] = []
    catalog = provider_catalog()
    search_provider_id = args.search_provider
    extract_provider_id = args.extract_provider
    effective_extract_provider_id = "none" if args.no_crawl_pages else extract_provider_id
    prompt_fn = getpass.getpass if args.prompt_for_keys else None

    search_api_key: str | None = None
    if args.fixture:
        fixture_results = read_fixture(args.fixture)
        raw_results.extend(fixture_results)
        sources.append(SearchSource("fixture", len(fixture_results), args.fixture))
        effective_search_provider = "fixture"
    else:
        if search_provider_id == "codex_manual":
            raise SystemExit("Use --manual-seeds or --fixture with --search-provider codex_manual.")
        deprecated_serpapi_key = args.serpapi_key if search_provider_id == "serpapi" else None
        search_api_key = resolve_provider_key(
            search_provider_id,
            catalog["search"][search_provider_id],
            args.search_api_key or deprecated_serpapi_key,
            prompt_fn=prompt_fn,
        )
        if catalog["search"][search_provider_id].get("requires_key") and not search_api_key:
            env_name = catalog["search"][search_provider_id]["env"]
            raise SystemExit(f"Set {env_name}, pass --search-api-key, or use --prompt-for-keys.")
        per_query = max(1, args.max_results // max(1, len(queries)))
        for query in queries:
            try:
                results = search_provider(query, search_provider_id, search_api_key or "", per_query)
            except Exception as exc:  # noqa: BLE001 - one failed query should not lose the whole run
                safe_exc = _redact_key(str(exc), search_api_key or "")
                print(f"[warn] query failed ({safe_exc}); skipping: {query!r}", file=sys.stderr)
                sources.append(SearchSource(query, 0, f"error:{safe_exc}"))
                continue
            raw_results.extend(results)
            sources.append(SearchSource(query, len(results), search_provider_id))
        effective_search_provider = search_provider_id

    manual_results: list[dict[str, Any]] = []
    if args.manual_seeds:
        manual_results = read_manual_seeds(args.manual_seeds)
        raw_results.extend(manual_results)
        sources.append(SearchSource("manual seeds", len(manual_results), args.manual_seeds))

    for item in raw_results:
        url = item.get("link") or item.get("url") or ""
        if url and is_blocked_url(url):
            rejected.append({"url": url, "reason": "blocked source; use as manual reference only"})

    leads = leads_from_search_results(raw_results, " | ".join(queries), theme_id, theme)
    merge_linkedin_references(leads, manual_results)
    leads = leads[: args.max_results]

    extract_api_key = resolve_provider_key(
        effective_extract_provider_id,
        catalog["extract"][effective_extract_provider_id],
        args.extract_api_key,
        prompt_fn=prompt_fn,
    )
    if catalog["extract"][effective_extract_provider_id].get("requires_key") and not extract_api_key:
        env_name = catalog["extract"][effective_extract_provider_id]["env"]
        raise SystemExit(f"Set {env_name}, pass --extract-api-key, or use --prompt-for-keys.")
    enrich_public_pages(leads, theme, effective_extract_provider_id, extract_api_key)
    if args.contact_search and effective_search_provider not in ("fixture", "codex_manual"):
        sources.extend(
            enrich_contacts_via_search(
                leads,
                search_provider_id,
                search_api_key or "",
                theme,
                per_lead_queries=args.contact_search_queries,
                budget=args.contact_search_budget,
            )
        )

    config = {
        "theme_id": theme_id,
        "theme_label": theme.get("label", theme_id),
        "location": args.location,
        "search_provider": effective_search_provider,
        "extract_provider": extract_provider_id,
        "effective_extract_provider": effective_extract_provider_id,
        "search_provider_env": catalog["search"].get(search_provider_id, {}).get("env", ""),
        "extract_provider_env": catalog["extract"].get(extract_provider_id, {}).get("env", ""),
        "contact_search": args.contact_search,
        "contact_search_queries": args.contact_search_queries,
        "contact_search_budget": args.contact_search_budget,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "queries": queries,
        "linkedin_guardrail": prebuilt_themes()["linkedin-assisted-cross-reference"]["guardrail"],
    }
    export_workbook(leads, sources, rejected, config, args.output)
    return Path(args.output)


def main() -> int:
    args = parse_args()
    output = run(args)
    if not args.list_themes and not args.list_providers:
        print(f"Wrote {output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
